from __future__ import annotations

import csv
import json
import logging
import os
import platform
import re
import sys
import unicodedata
from calendar import monthrange
from datetime import datetime, timedelta
from pathlib import Path

from PySide6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QFrame, QStackedWidget, QGraphicsDropShadowEffect, QMessageBox, QFileDialog,
    QDialog, QLineEdit, QDialogButtonBox, QComboBox, QProgressDialog
)
from PySide6.QtCore import Qt, QTimer, QObject, Signal, QRunnable, QThreadPool, Slot, QStandardPaths
from PySide6.QtGui import QColor, QKeySequence, QShortcut, QGuiApplication

import database.db as db
import controllers.empresa_controller as empresa_controller
import controllers.pagamento_controller as pagamento_controller
from services import email_service
from services.dashboard_ops_service import (
    build_jobs_status as _build_jobs_status_payload,
    build_operational_summary_text as _build_operational_summary_payload,
)
from services import finance_payload_service
from services import dashboard_payload_service
from services import clientes_service
from models.activity_models import ActivityEntry

from views.login_view import LoginView
from views.dashboard_view import DashboardView
from views.financeiro_view import FinanceiroView
from views.cadastro_cliente_view import CadastroClienteView
from views.cadastro_empresa_view import CadastroEmpresaView
from views.registrar_pagamento_view import RegistrarPagamentoView
from views.listar_clientes_view import ListarClientesView
from views.listar_empresas_view import ListarEmpresasView
from views.relatorios_view import RelatoriosView
from views.role_utils import normalize_role as _shared_normalize_role
from openpyxl import Workbook, load_workbook   
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def _env_flag(name: str, default: bool = False) -> bool:
    raw = os.getenv(name)
    if raw is None:
        return bool(default)
    return str(raw).strip().lower() in {"1", "true", "yes", "on"}


def _sanitize_error_text(text: str) -> str:
    msg = str(text or "")
    msg = re.sub(r"(postgres(?:ql)?://)([^\\s]+)", r"\1***", msg, flags=re.IGNORECASE)
    msg = re.sub(r"(?i)(password\\s*=\\s*)([^\\s;]+)", r"\1***", msg)
    msg = re.sub(r"(?i)(pwd\\s*=\\s*)([^\\s;]+)", r"\1***", msg)
    msg = re.sub(r"(?i)(token\\s*=\\s*)([^\\s;]+)", r"\1***", msg)
    msg = re.sub(r"(?i)(apikey\\s*=\\s*)([^\\s;]+)", r"\1***", msg)
    return msg.strip()


def _only_digits(value: str) -> str:
    return "".join(ch for ch in str(value or "") if ch.isdigit())


def _is_windows_11() -> bool:
    if sys.platform != "win32":
        return False
    try:
        build = int(str(platform.version()).split(".")[-1])
        return build >= 22000
    except Exception:
        return False


ROLE_ADMIN = "admin"
ROLE_FUNCIONARIO = "funcionario"
ROLE_RECEPCAO = "recepcao"


def _normalize_role(nivel: str) -> str:
    plain = _shared_normalize_role(nivel)
    if plain in {"admin", "administrador"} or "admin" in plain:
        return ROLE_ADMIN
    if plain in {"funcionario"} or plain.startswith("func"):
        return ROLE_FUNCIONARIO
    if plain in {"recepcao", "recepcionista"} or plain.startswith("recep"):
        return ROLE_RECEPCAO
    return plain


# ============================
# Helpers de mÃªs
# ============================
_PT_BR_MONTHS = {
    "JAN": "01", "FEV": "02", "MAR": "03", "ABR": "04", "MAI": "05", "JUN": "06",
    "JUL": "07", "AGO": "08", "SET": "09", "OUT": "10", "NOV": "11", "DEZ": "12",
}
_NUM_TO_PT = {v: k for k, v in _PT_BR_MONTHS.items()}


def mes_ref_to_iso(mes_ref: str) -> str:
    s = (mes_ref or "").strip().upper()

    if len(s) == 7 and s[4] == "-":
        y, m = s.split("-")
        if len(y) == 4 and m.isdigit() and 1 <= int(m) <= 12:
            return f"{y}-{int(m):02d}"
        raise ValueError("MÃªs ISO invÃ¡lido")

    if "/" in s:
        mm, yy = s.split("/", 1)
        mm = mm.strip().upper()
        yy = yy.strip()
        if mm in _PT_BR_MONTHS and yy.isdigit() and len(yy) == 4:
            return f"{yy}-{_PT_BR_MONTHS[mm]}"
        raise ValueError("MÃªs BR invÃ¡lido")

    raise ValueError("Formato de mÃªs invÃ¡lido")


def iso_to_mes_ref_br(yyyy_mm: str) -> str:
    s = (yyyy_mm or "").strip()
    if len(s) == 7 and s[4] == "-":
        y, m = s.split("-")
        m2 = f"{int(m):02d}"
        return f"{_NUM_TO_PT.get(m2, m2)}/{y}"
    return s


# ============================
# Worker infra
# ============================
class _WorkerSignals(QObject):
    result = Signal(object)
    error = Signal(str)


class _Worker(QRunnable):
    def __init__(self, fn, *args, **kwargs):
        super().__init__()
        self.fn = fn
        self.args = args
        self.kwargs = kwargs
        self.signals = _WorkerSignals()

    @Slot()
    def run(self):
        try:
            out = self.fn(*self.args, **self.kwargs)
            self.signals.result.emit(out)
        except Exception as e:
            self.signals.error.emit(_sanitize_error_text(str(e)))


# ============================
# TitleBar
# ============================
class TitleBar(QFrame):
    def __init__(self, parent_window: QMainWindow):
        super().__init__()
        self.parent_window = parent_window
        self.setFixedHeight(44)
        self.setObjectName("titleBar")
        self._drag_pos = None

        layout = QHBoxLayout(self)
        layout.setContentsMargins(14, 0, 10, 0)
        layout.setSpacing(10)

        self.title_label = QLabel("MedContract")
        self.title_label.setObjectName("titleText")

        layout.addWidget(self.title_label)
        layout.addStretch()

        self.btn_min = QPushButton("—")
        self.btn_min.setObjectName("winMin")

        self.btn_close = QPushButton("✕")
        self.btn_close.setObjectName("winClose")

        self.btn_min.clicked.connect(self.parent_window.showMinimized)
        self.btn_close.clicked.connect(self.parent_window.close)

        layout.addWidget(self.btn_min)
        layout.addWidget(self.btn_close)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self._drag_pos = event.globalPosition().toPoint() - self.parent_window.frameGeometry().topLeft()
            event.accept()

    def mouseMoveEvent(self, event):
        if event.buttons() == Qt.LeftButton and self._drag_pos is not None:
            self.parent_window.move(event.globalPosition().toPoint() - self._drag_pos)
            event.accept()

    def mouseReleaseEvent(self, event):
        self._drag_pos = None
        event.accept()


# ============================
# Dialog export pagamentos mÃªs
# ============================
class ExportPagamentosDialog(QDialog):
    def __init__(self, parent=None, default_iso: str = ""):
        super().__init__(parent)
        self.setWindowTitle("Exportar pagamentos do mÃªs")
        self.setModal(True)
        self.setObjectName("ExportDlg")
        self.setFixedWidth(460)

        root = QVBoxLayout(self)
        root.setContentsMargins(18, 16, 18, 14)
        root.setSpacing(12)

        title = QLabel("Exportar pagamentos do mÃªs")
        title.setObjectName("dlgTitle")

        sub = QLabel("Informe o mÃªs para exportar (AAAA-MM ou JAN/AAAA).")
        sub.setObjectName("dlgSub")
        sub.setWordWrap(True)

        box = QFrame()
        box.setObjectName("dlgBox")
        box_l = QVBoxLayout(box)
        box_l.setContentsMargins(14, 14, 14, 14)
        box_l.setSpacing(8)

        self.input_mes = QLineEdit()
        self.input_mes.setObjectName("dlgInput")
        self.input_mes.setPlaceholderText("Ex: 2026-02 ou FEV/2026")
        self.input_mes.setFixedHeight(40)
        self.input_mes.setText(default_iso or "")

        hint = QLabel("Dica: use JAN/2026, FEV/2026â€¦ ou 2026-01, 2026-02â€¦")
        hint.setObjectName("dlgHint")
        hint.setWordWrap(True)

        self.msg = QLabel("")
        self.msg.setObjectName("dlgMsg")
        self.msg.setVisible(False)

        lab = QLabel("MÃªs:")
        lab.setObjectName("dlgLabel")

        box_l.addWidget(lab)
        box_l.addWidget(self.input_mes)
        box_l.addWidget(hint)
        box_l.addWidget(self.msg)

        btns = QDialogButtonBox(QDialogButtonBox.Cancel | QDialogButtonBox.Ok)
        btns.setObjectName("dlgBtns")
        btns.button(QDialogButtonBox.Ok).setText("Exportar")
        btns.button(QDialogButtonBox.Cancel).setText("Cancelar")

        btns.accepted.connect(self._on_ok)
        btns.rejected.connect(self.reject)

        root.addWidget(title)
        root.addWidget(sub)
        root.addWidget(box)
        root.addWidget(btns)

        self._mes_iso = None
        self._apply_styles()

        self.input_mes.setFocus()
        self.input_mes.selectAll()

    def _apply_styles(self):
        self.setStyleSheet("""
        QDialog#ExportDlg { background: #f8fafc; font-family: Segoe UI; border-radius: 14px; }

        QLabel#dlgTitle { font-size: 15px; font-weight: 900; color: #0f172a; }
        QLabel#dlgSub { font-size: 12px; font-weight: 700; color: #64748b; }
        QLabel#dlgLabel { font-size: 12px; font-weight: 900; color: #0f172a; }

        QFrame#dlgBox {
            background: rgba(255,255,255,0.92);
            border: 1px solid rgba(15, 23, 42, 0.10);
            border-radius: 14px;
        }

        QLineEdit#dlgInput {
            border: 1px solid rgba(15, 23, 42, 0.16);
            border-radius: 12px;
            padding-left: 12px;
            font-size: 13px;
            background: white;
            color: #0f172a;
        }
        QLineEdit#dlgInput:focus { border: 1px solid rgba(43,108,126,0.72); }

        QLabel#dlgHint { font-size: 11px; color: #94a3b8; font-weight: 700; }
        QLabel#dlgMsg {
            background: rgba(231, 76, 60, 0.10);
            border: 1px solid rgba(231, 76, 60, 0.25);
            color: #c0392b;
            padding: 8px 10px;
            border-radius: 12px;
            font-size: 12px;
            font-weight: 800;
        }

        QDialogButtonBox#dlgBtns QPushButton {
            min-height: 36px;
            border-radius: 12px;
            padding: 6px 12px;
            font-weight: 900;
        }
        QDialogButtonBox#dlgBtns QPushButton[text="Cancelar"] {
            background: rgba(255,255,255,0.95);
            border: 1px solid rgba(15, 23, 42, 0.12);
            color: #0f172a;
        }
        QDialogButtonBox#dlgBtns QPushButton[text="Exportar"] {
            background: #2b6c7e;
            border: 1px solid rgba(255,255,255,0.22);
            color: white;
        }
        QDialogButtonBox#dlgBtns QPushButton[text="Exportar"]:hover { background: #2f768a; }
        """)

    def _error(self, text: str):
        self.msg.setText(text)
        self.msg.setVisible(True)

    def _on_ok(self):
        raw = (self.input_mes.text() or "").strip()
        if not raw:
            self._error("Informe o mÃªs.")
            return
        try:
            self._mes_iso = mes_ref_to_iso(raw)
        except Exception:
            self._error("MÃªs invÃ¡lido. Use AAAA-MM ou JAN/AAAA.")
            return
        self.accept()

    def mes_iso(self) -> str:
        return self._mes_iso


class ContractTypeDialog(QDialog):
    def __init__(self, parent=None, default_type: str = "boleto"):
        super().__init__(parent)
        self.setWindowTitle("Gerar contrato em PDF")
        self.setModal(True)
        self.setObjectName("ContractTypeDlg")
        self.setFixedWidth(430)

        root = QVBoxLayout(self)
        root.setContentsMargins(18, 16, 18, 14)
        root.setSpacing(12)

        title = QLabel("Gerar contrato em PDF")
        title.setObjectName("dlgTitle")

        sub = QLabel("Escolha o tipo de contrato e a pasta de destino para gerar e baixar o PDF.")
        sub.setObjectName("dlgSub")
        sub.setWordWrap(True)

        box = QFrame()
        box.setObjectName("dlgBox")
        box_l = QVBoxLayout(box)
        box_l.setContentsMargins(14, 14, 14, 14)
        box_l.setSpacing(8)

        lab = QLabel("Tipo de contrato:")
        lab.setObjectName("dlgLabel")

        self.combo = QComboBox()
        self.combo.setObjectName("dlgCombo")
        self.combo.setFixedHeight(40)
        self.combo.addItem("Pix", "pix")
        self.combo.addItem("Boleto", "boleto")
        self.combo.addItem("Recepção", "recepcao")

        idx = max(0, self.combo.findData((default_type or "boleto").strip().lower()))
        self.combo.setCurrentIndex(idx)

        destino_lab = QLabel("Salvar em:")
        destino_lab.setObjectName("dlgLabel")

        self.dest_combo = QComboBox()
        self.dest_combo.setObjectName("dlgCombo")
        self.dest_combo.setFixedHeight(40)
        self._custom_output_dir = ""
        self._selected_output_dir = ""
        self._populate_destination_options()
        self.dest_combo.currentIndexChanged.connect(self._on_destination_changed)

        self.dest_hint = QLabel("")
        self.dest_hint.setObjectName("dlgSub")
        self.dest_hint.setWordWrap(True)
        self._set_destination_hint(self._selected_output_dir)

        self.msg = QLabel("")
        self.msg.setObjectName("dlgMsg")
        self.msg.setVisible(False)

        box_l.addWidget(lab)
        box_l.addWidget(self.combo)
        box_l.addWidget(destino_lab)
        box_l.addWidget(self.dest_combo)
        box_l.addWidget(self.dest_hint)
        box_l.addWidget(self.msg)

        btns = QDialogButtonBox(QDialogButtonBox.Cancel | QDialogButtonBox.Ok)
        btns.setObjectName("dlgBtns")
        btns.button(QDialogButtonBox.Ok).setText("Gerar PDF")
        btns.button(QDialogButtonBox.Cancel).setText("Cancelar")
        btns.accepted.connect(self._on_ok)
        btns.rejected.connect(self.reject)

        root.addWidget(title)
        root.addWidget(sub)
        root.addWidget(box)
        root.addWidget(btns)

        self._selected = None
        self._apply_styles()

    def _preferred_user_dir(self, standard_location, fallback: Path) -> str:
        raw = str(QStandardPaths.writableLocation(standard_location) or "").strip()
        if raw:
            return str(Path(raw))
        return str(fallback)

    def _populate_destination_options(self):
        options: list[tuple[str, str]] = []
        seen: set[str] = set()

        def _add(label: str, path_txt: str):
            path_norm = str(Path(path_txt or "").expanduser())
            if not path_norm:
                return
            key = path_norm.lower()
            if key in seen:
                return
            seen.add(key)
            options.append((label, path_norm))

        _add(
            "Downloads (padrão)",
            self._preferred_user_dir(QStandardPaths.DownloadLocation, Path.home() / "Downloads"),
        )
        _add(
            "Área de trabalho",
            self._preferred_user_dir(QStandardPaths.DesktopLocation, Path.home() / "Desktop"),
        )
        _add(
            "Documentos",
            self._preferred_user_dir(QStandardPaths.DocumentsLocation, Path.home() / "Documents"),
        )

        for label, path_txt in options:
            self.dest_combo.addItem(label, path_txt)
        self.dest_combo.addItem("Escolher pasta...", "__custom__")

        self.dest_combo.setCurrentIndex(0)
        self._selected_output_dir = str(self.dest_combo.itemData(0) or "").strip()

    def _set_destination_hint(self, path_txt: str):
        dest = str(path_txt or "").strip()
        if not dest:
            self.dest_hint.setText("Selecione uma pasta de destino.")
            return
        self.dest_hint.setText(f"Pasta de destino: {dest}")

    def _on_destination_changed(self, _index: int):
        data = str(self.dest_combo.currentData() or "").strip()
        if data == "__custom__":
            start_dir = self._custom_output_dir or self._selected_output_dir or str(Path.home())
            chosen = QFileDialog.getExistingDirectory(
                self,
                "Escolher pasta para salvar o PDF do contrato",
                start_dir,
            )
            if chosen:
                self._custom_output_dir = str(Path(chosen))
                self._selected_output_dir = self._custom_output_dir
                self._set_destination_hint(self._selected_output_dir)
            else:
                self.dest_combo.blockSignals(True)
                self.dest_combo.setCurrentIndex(0)
                self.dest_combo.blockSignals(False)
                self._selected_output_dir = str(self.dest_combo.itemData(0) or "").strip()
                self._set_destination_hint(self._selected_output_dir)
            return

        self._selected_output_dir = data
        self._set_destination_hint(self._selected_output_dir)

    def _apply_styles(self):
        self.setStyleSheet("""
        QDialog#ContractTypeDlg { background: #f8fafc; font-family: Segoe UI; border-radius: 14px; }
        QLabel#dlgTitle { font-size: 15px; font-weight: 900; color: #0f172a; }
        QLabel#dlgSub { font-size: 12px; font-weight: 700; color: #64748b; }
        QLabel#dlgLabel { font-size: 12px; font-weight: 900; color: #0f172a; }
        QFrame#dlgBox {
            background: rgba(255,255,255,0.92);
            border: 1px solid rgba(15, 23, 42, 0.10);
            border-radius: 14px;
        }
        QComboBox#dlgCombo {
            border: 1px solid rgba(15, 23, 42, 0.16);
            border-radius: 12px;
            padding: 0 36px 0 12px;
            font-size: 13px;
            background: white;
            color: #0f172a;
        }
        QComboBox#dlgCombo:hover { border-color: rgba(15, 23, 42, 0.24); }
        QComboBox#dlgCombo:focus { border: 1px solid rgba(43,108,126,0.72); }
        QComboBox#dlgCombo::drop-down { border: none; width: 28px; }
        QComboBox#dlgCombo::down-arrow {
            width: 0; height: 0;
            border-left: 4px solid transparent;
            border-right: 4px solid transparent;
            border-top: 5px solid #64748b;
            margin-right: 6px;
        }
        QComboBox QAbstractItemView {
            background: white;
            border: 1px solid rgba(15, 23, 42, 0.12);
            selection-background-color: rgba(43,108,126,0.12);
            padding: 4px;
        }
        QLabel#dlgMsg {
            background: rgba(231, 76, 60, 0.10);
            border: 1px solid rgba(231, 76, 60, 0.25);
            color: #c0392b;
            padding: 8px 10px;
            border-radius: 12px;
            font-size: 12px;
            font-weight: 800;
        }
        QDialogButtonBox#dlgBtns QPushButton {
            min-height: 36px;
            border-radius: 12px;
            padding: 6px 12px;
            font-weight: 900;
        }
        QDialogButtonBox#dlgBtns QPushButton[text="Cancelar"] {
            background: rgba(255,255,255,0.95);
            border: 1px solid rgba(15, 23, 42, 0.12);
            color: #0f172a;
        }
        QDialogButtonBox#dlgBtns QPushButton[text="Gerar PDF"] {
            background: #2b6c7e;
            border: 1px solid rgba(255,255,255,0.22);
            color: white;
        }
        QDialogButtonBox#dlgBtns QPushButton[text="Gerar PDF"]:hover { background: #2f768a; }
        """)

    def _on_ok(self):
        selected = self.combo.currentData()
        if not selected:
            self.msg.setText("Selecione um tipo de contrato.")
            self.msg.setVisible(True)
            return
        if not str(self._selected_output_dir or "").strip():
            self.msg.setText("Selecione uma pasta para salvar o contrato.")
            self.msg.setVisible(True)
            return
        self._selected = str(selected)
        self.accept()

    def selected_type(self) -> str | None:
        return self._selected

    def selected_output_dir(self) -> str | None:
        dest = str(self._selected_output_dir or "").strip()
        return dest or None


# ============================
# MainWindow
# ============================
class MainWindow(QMainWindow):
    APP_VERSION = "2.0.0"

    def __init__(self):
        super().__init__()

        self._performance_mode = _env_flag("MEDCONTRACT_PERF_MODE", default=_is_windows_11())
        self.setWindowFlags(Qt.FramelessWindowHint)
        # Em notebooks com GPU integrada (comum no Win11 corporativo),
        # transparência do topo + sombra do contêiner pode causar "lag" visual.
        self.setAttribute(Qt.WA_TranslucentBackground, not self._performance_mode)

        self._thread_pool = QThreadPool.globalInstance()
        self._refresh_inflight = False
        self._finance_inflight = False
        self._contas_inflight = False
        self._dashboard_refresh_pending = False
        self._dashboard_refresh_pending_force = False
        self._finance_refresh_pending = False
        self._finance_refresh_pending_ref = ""
        self._finance_refresh_pending_force = False
        self._finance_refresh_pending_query: dict | None = None
        self._contas_refresh_pending = False
        self._contas_refresh_pending_ref = ""
        self._contas_refresh_pending_force = False
        self._contas_refresh_pending_query: dict | None = None
        self._dashboard_period = "month"
        try:
            self._dashboard_cache_ttl_s = max(
                0,
                int((os.getenv("MEDCONTRACT_DASHBOARD_CACHE_TTL") or "15").strip()),
            )
        except Exception:
            self._dashboard_cache_ttl_s = 15
        try:
            self._finance_cache_ttl_s = max(
                0,
                int((os.getenv("MEDCONTRACT_FINANCEIRO_CACHE_TTL") or "10").strip()),
            )
        except Exception:
            self._finance_cache_ttl_s = 10
        self._dashboard_cache: dict[str, tuple[datetime, dict]] = {}
        self._finance_cache: dict[str, tuple[datetime, dict]] = {}
        self._contas_cache: dict[str, tuple[datetime, dict]] = {}
        self._export_history: list[dict] = []
        self._activity_history: list[dict] = []
        self._last_operational_summary_date = ""
        self._last_due_digest_date = ""
        self._email_workers: list[_Worker] = []
        self._cobranca_workers: list[_Worker] = []
        self._contract_workers: list[_Worker] = []
        self._cliente_save_workers: list[_Worker] = []
        self._pagamento_register_workers: list[_Worker] = []
        self._cliente_delete_workers: list[_Worker] = []
        self._empresa_delete_workers: list[_Worker] = []
        self._cancelar_plano_workers: list[_Worker] = []
        self._renovar_contrato_workers: list[_Worker] = []
        self._renovar_lote_workers: list[_Worker] = []
        self._reajuste_workers: list[_Worker] = []
        self._status_sync_workers: list[_Worker] = []
        self._cliente_save_inflight = False
        self._pagamento_register_inflight = False
        self._cliente_delete_inflight = False
        self._empresa_delete_inflight = False
        self._cancelar_plano_inflight = False
        self._renovar_contrato_inflight = False
        self._renovar_lote_inflight = False
        self._reajuste_inflight = False
        self._status_sync_inflight = False
        self._status_sync_pending_force = False
        self._empresa_save_workers: list[_Worker] = []
        self._empresa_save_inflight = False
        self._nivel_usuario = ""
        self._usuario_atual = ""
        self._layout_density = "normal"
        self._finance_prefs_loaded_user = ""
        self._finance_pref_query: dict = {}
        self._contas_pref_query: dict = {}
        self._last_auto_export_key = ""
        self._last_auto_cobranca_key = ""
        self._auto_cobranca_inflight = False
        self._auto_export_timer = QTimer(self)
        self._auto_export_timer.setInterval(600_000)
        self._auto_export_timer.timeout.connect(self._auto_export_tick)
        if _env_flag("MEDCONTRACT_AUTO_EXPORT_ENABLED", False):
            self._auto_export_timer.start()
            QTimer.singleShot(5_000, self._auto_export_tick)

        self._auto_cobranca_timer = QTimer(self)
        self._auto_cobranca_timer.setInterval(600_000)
        self._auto_cobranca_timer.timeout.connect(self._auto_cobranca_tick)
        if _env_flag("MEDCONTRACT_AUTO_COBRANCA_ENABLED", True):
            self._auto_cobranca_timer.start()
            QTimer.singleShot(8_000, self._auto_cobranca_tick)

        self._apply_best_window_size()

        root = QWidget()
        root.setObjectName("windowRoot")
        root.setAttribute(Qt.WA_StyledBackground, True)
        root_layout = QVBoxLayout(root)
        self._root_layout = root_layout
        if self._performance_mode:
            root_layout.setContentsMargins(0, 0, 0, 0)
        else:
            root_layout.setContentsMargins(18, 18, 18, 18)
        root_layout.setSpacing(0)

        self.container = QFrame()
        self.container.setObjectName("appContainer")

        if not self._performance_mode:
            shadow = QGraphicsDropShadowEffect(self.container)
            shadow.setBlurRadius(30)
            shadow.setOffset(0, 8)
            shadow.setColor(QColor(0, 0, 0, 70))
            self.container.setGraphicsEffect(shadow)

        container_layout = QVBoxLayout(self.container)
        container_layout.setContentsMargins(0, 0, 0, 0)
        container_layout.setSpacing(0)

        self.titlebar = TitleBar(self)
        container_layout.addWidget(self.titlebar)

        self.stack = QStackedWidget()
        self.stack.setObjectName("stackArea")
        container_layout.addWidget(self.stack)

        root_layout.addWidget(self.container)
        self.setCentralWidget(root)

        self.setToolTip("")
        root.setToolTip("")
        self.container.setToolTip("")
        self.stack.setToolTip("")
        self.titlebar.setToolTip("")
        self.titlebar.title_label.setToolTip("")
        self.titlebar.btn_min.setToolTip("")
        self.titlebar.btn_close.setToolTip("")

        self.apply_styles()

        db.create_tables()
        if hasattr(db, "create_default_users"):
            db.create_default_users(required_if_empty=False)
        else:
            db.create_default_admin()
            if hasattr(db, "create_default_recepcao"):
                db.create_default_recepcao()

        self.login = LoginView()
        self.dashboard = DashboardView()
        self.financeiro = FinanceiroView()
        self.cadastro = CadastroClienteView()
        self.cadastro_empresa = CadastroEmpresaView()
        self.pagamento = RegistrarPagamentoView()
        self.listar = ListarClientesView()
        self.listar_empresas = ListarEmpresasView()
        self.relatorios = RelatoriosView()

        self.stack.addWidget(self.login)
        self.stack.addWidget(self.dashboard)
        self.stack.addWidget(self.financeiro)
        self.stack.addWidget(self.cadastro)
        self.stack.addWidget(self.cadastro_empresa)
        self.stack.addWidget(self.pagamento)
        self.stack.addWidget(self.listar)
        self.stack.addWidget(self.listar_empresas)
        self.stack.addWidget(self.relatorios)

        self.stack.setCurrentWidget(self.login)

        self.conectar_navegacao()

        self._shortcut_backup = QShortcut(QKeySequence("Ctrl+B"), self)
        self._shortcut_backup.activated.connect(self.fazer_backup)
        self._shortcut_backup.setEnabled(False)

        self._shortcut_refresh = QShortcut(QKeySequence("F5"), self)
        self._shortcut_refresh.activated.connect(self._on_global_refresh)

        self._finance_prefs_save_timer = QTimer(self)
        self._finance_prefs_save_timer.setSingleShot(True)
        self._finance_prefs_save_timer.setInterval(420)
        self._finance_prefs_save_timer.timeout.connect(self._persist_finance_preferences)

        self._auto_refresh_timer = QTimer(self)
        self._auto_refresh_timer.setInterval(180_000 if self._performance_mode else 60_000)
        self._auto_refresh_timer.timeout.connect(self._auto_refresh_tick)
        self._auto_refresh_timer.start()

        self._status_sync_timer = QTimer(self)
        self._status_sync_timer.setSingleShot(True)
        self._status_sync_timer.timeout.connect(self._on_status_sync_daily_tick)
        self._schedule_next_status_sync()
        QTimer.singleShot(2_000, lambda: self._run_status_sync_async(force=True, reason="startup"))

    def _role(self) -> str:
        return _normalize_role(self._nivel_usuario)

    def _show_modern_question(
        self,
        *,
        title: str,
        subtitle: str = "",
        details: str = "",
        confirm_text: str = "Confirmar",
        cancel_text: str = "Cancelar",
    ) -> bool:
        dlg = QDialog(self)
        dlg.setModal(True)
        dlg.setWindowTitle(title)
        dlg.setObjectName("SaasConfirmDialog")
        dlg.setMinimumWidth(540)

        root = QVBoxLayout(dlg)
        root.setContentsMargins(20, 18, 20, 16)
        root.setSpacing(10)

        lbl_title = QLabel(title)
        lbl_title.setObjectName("dlgTitle")
        root.addWidget(lbl_title)

        if str(subtitle or "").strip():
            lbl_sub = QLabel(str(subtitle or "").strip())
            lbl_sub.setObjectName("dlgSub")
            lbl_sub.setWordWrap(True)
            root.addWidget(lbl_sub)

        if str(details or "").strip():
            lbl_details = QLabel(str(details or "").strip())
            lbl_details.setObjectName("dlgBody")
            lbl_details.setWordWrap(True)
            root.addWidget(lbl_details)

        actions = QHBoxLayout()
        actions.addStretch()
        btn_cancel = QPushButton(cancel_text)
        btn_cancel.setObjectName("dlgBtnSecondary")
        btn_cancel.setFixedHeight(36)
        btn_ok = QPushButton(confirm_text)
        btn_ok.setObjectName("dlgBtnPrimary")
        btn_ok.setFixedHeight(36)
        actions.addWidget(btn_cancel)
        actions.addWidget(btn_ok)
        root.addLayout(actions)

        dlg.setStyleSheet(
            """
            QDialog#SaasConfirmDialog {
                background: #ffffff;
                border: 1px solid rgba(15, 23, 42, 0.12);
                border-radius: 14px;
                font-family: Segoe UI;
            }
            QLabel#dlgTitle {
                font-size: 18px;
                font-weight: 800;
                color: #0f172a;
            }
            QLabel#dlgSub {
                font-size: 12px;
                font-weight: 600;
                color: #475569;
            }
            QLabel#dlgBody {
                background: #f8fafc;
                border: 1px solid #e2e8f0;
                border-radius: 10px;
                padding: 12px 14px;
                color: #0f172a;
                font-size: 12px;
            }
            QPushButton#dlgBtnPrimary {
                min-width: 160px;
                border: none;
                border-radius: 8px;
                background: #2b6c7e;
                color: white;
                font-weight: 800;
                padding: 0 14px;
            }
            QPushButton#dlgBtnPrimary:hover { background: #2f768a; }
            QPushButton#dlgBtnSecondary {
                border: 1px solid rgba(15, 23, 42, 0.16);
                border-radius: 8px;
                background: #ffffff;
                color: #0f172a;
                font-weight: 700;
                padding: 0 14px;
            }
            QPushButton#dlgBtnSecondary:hover {
                border-color: rgba(43, 108, 126, 0.62);
                color: #1f5f72;
            }
            """
        )

        btn_ok.clicked.connect(dlg.accept)
        btn_cancel.clicked.connect(dlg.reject)
        return dlg.exec() == QDialog.Accepted

    def _can_access_financeiro(self) -> bool:
        role = self._role()
        return role not in {ROLE_FUNCIONARIO, ROLE_RECEPCAO}

    def _can_export(self) -> bool:
        role = self._role()
        return bool(role) and role != ROLE_RECEPCAO

    def _can_backup(self) -> bool:
        return self._role() == ROLE_ADMIN

    def _can_edit_cliente(self) -> bool:
        return self._role() != ROLE_RECEPCAO

    def _notify_access_denied(self, message: str, *, popup: bool = False):
        msg = (message or "Acesso restrito para este perfil.").strip()
        try:
            if hasattr(self.dashboard, "show_error"):
                self.dashboard.show_error(msg)
        except Exception:
            pass
        if popup:
            QMessageBox.warning(self, "Acesso restrito", msg)

    def _ensure_export_allowed(self, *, popup: bool = True) -> bool:
        if not self._role():
            self._notify_access_denied(
                "Faça login para exportar dados.",
                popup=popup,
            )
            return False
        if self._can_export():
            return True
        self._notify_access_denied(
            "Perfil de recepção não pode exportar dados.",
            popup=popup,
        )
        return False

    # ============================
    # Window sizing
    # ============================
    def _apply_best_window_size(self):
        self.setMinimumSize(1280, 820)

        screen = QGuiApplication.primaryScreen()
        if not screen:
            self.resize(1500, 920)
            return

        geo = screen.availableGeometry()
        target_w = int(geo.width() * 0.92)
        target_h = int(geo.height() * 0.92)

        target_w = max(1280, min(target_w, 1760))
        target_h = max(820, min(target_h, 980))

        self.resize(target_w, target_h)

        x = geo.x() + (geo.width() - target_w) // 2
        y = geo.y() + (geo.height() - target_h) // 2
        self.move(x, y)

    # ============================
    # Styles
    # ============================
    def apply_styles(self):
        root_bg = "#f4f6f9" if self._performance_mode else "transparent"
        radius = "0px" if self._performance_mode else "18px"
        title_bg = "#ffffff" if self._performance_mode else "rgba(255,255,255,0.75)"
        stack_bg = "#f4f6f9" if self._performance_mode else "transparent"

        self.setStyleSheet(f"""
        QWidget#windowRoot {{ background-color: {root_bg}; }}
        QFrame#appContainer {{ background-color: #f4f6f9; border-radius: {radius}; }}
        QFrame#titleBar {{
            background-color: {title_bg};
            border-top-left-radius: {radius};
            border-top-right-radius: {radius};
            border-bottom: 1px solid #e8eaed;
        }}
        QLabel#titleText {{ font-size: 14px; font-weight: 700; color: #0f172a; }}
        QPushButton#winMin, QPushButton#winClose {{
            width: 36px; height: 28px; border: none; border-radius: 8px;
            font-size: 14px; font-weight: 900; background: transparent; color: #0f172a;
        }}
        QPushButton#winMin:hover {{ background-color: rgba(37, 99, 235, 0.12); }}
        QPushButton#winClose:hover {{ background-color: rgba(231, 76, 60, 0.16); color: #c0392b; }}
        QStackedWidget#stackArea {{
            background: {stack_bg};
            border-bottom-left-radius: {radius};
            border-bottom-right-radius: {radius};
        }}
        """)

    # ============================
    # Wiring
    # ============================
    def conectar_navegacao(self):
        self.login.login_success.connect(self.ir_para_dashboard)

        self.dashboard.ir_cadastro_signal.connect(self.ir_para_cadastro_create)
        if hasattr(self.dashboard, "ir_novo_contrato_signal"):
            self.dashboard.ir_novo_contrato_signal.connect(self.ir_para_novo_contrato)
        self.dashboard.ir_pagamento_signal.connect(self.ir_para_pagamento)
        self.dashboard.ir_listar_signal.connect(self.ir_para_listar)
        if hasattr(self.dashboard, "ir_cadastro_empresa_signal"):
            self.dashboard.ir_cadastro_empresa_signal.connect(self.ir_para_cadastro_empresa_create)
        if hasattr(self.dashboard, "ir_listar_empresas_signal"):
            self.dashboard.ir_listar_empresas_signal.connect(self.ir_para_listar_empresas)
        if hasattr(self.dashboard, "ir_relatorios_signal"):
            self.dashboard.ir_relatorios_signal.connect(self.ir_para_relatorios)
        if hasattr(self.dashboard, "ir_financeiro_signal"):
            self.dashboard.ir_financeiro_signal.connect(self.ir_para_financeiro)
        if hasattr(self.dashboard, "ir_listar_filtrado_signal"):
            self.dashboard.ir_listar_filtrado_signal.connect(self.ir_para_listar_filtrado)
        if hasattr(self.dashboard, "busca_global_signal"):
            self.dashboard.busca_global_signal.connect(self.executar_busca_global)
        self.dashboard.logout_signal.connect(self.ir_para_login)

        if hasattr(self.dashboard, "refresh_signal"):
            self.dashboard.refresh_signal.connect(lambda: self.atualizar_dashboard_async(force=True))
        if hasattr(self.dashboard, "period_changed_signal"):
            self.dashboard.period_changed_signal.connect(self._on_dashboard_period_changed)
        if hasattr(self.dashboard, "export_clientes_signal"):
            self.dashboard.export_clientes_signal.connect(self.exportar_clientes)
        if hasattr(self.dashboard, "export_inadimplentes_signal"):
            self.dashboard.export_inadimplentes_signal.connect(self.exportar_inadimplentes)
        if hasattr(self.dashboard, "export_pagamentos_mes_signal"):
            self.dashboard.export_pagamentos_mes_signal.connect(self.exportar_pagamentos_mes)
        if hasattr(self.dashboard, "backup_now_signal"):
            self.dashboard.backup_now_signal.connect(self.fazer_backup)

        self.cadastro.voltar_signal.connect(lambda: self.stack.setCurrentWidget(self.dashboard))
        self.cadastro_empresa.voltar_signal.connect(lambda: self.stack.setCurrentWidget(self.dashboard))
        if hasattr(self.cadastro_empresa, "cancelar_signal"):
            self.cadastro_empresa.cancelar_signal.connect(self.ir_para_listar_empresas)
        self.pagamento.voltar_signal.connect(lambda: self.stack.setCurrentWidget(self.dashboard))
        self.listar.voltar_signal.connect(lambda: self.stack.setCurrentWidget(self.dashboard))
        self.listar_empresas.voltar_signal.connect(lambda: self.stack.setCurrentWidget(self.dashboard))
        self.relatorios.voltar_signal.connect(lambda: self.stack.setCurrentWidget(self.dashboard))
        self.financeiro.voltar_signal.connect(lambda: self.stack.setCurrentWidget(self.dashboard))
        self.financeiro.refresh_signal.connect(
            lambda mes: self.atualizar_financeiro_async(
                mes,
                force=True,
                query=(self.financeiro.current_query() if hasattr(self.financeiro, "current_query") else None),
            )
        )
        if hasattr(self.financeiro, "contas_refresh_signal"):
            self.financeiro.contas_refresh_signal.connect(
                lambda mes: self.atualizar_contas_pagar_async(
                    mes,
                    force=True,
                    query=(self.financeiro.current_contas_query() if hasattr(self.financeiro, "current_contas_query") else None),
                )
            )
        if hasattr(self.financeiro, "query_changed_signal"):
            self.financeiro.query_changed_signal.connect(self._on_finance_query_changed)
        if hasattr(self.financeiro, "contas_query_changed_signal"):
            self.financeiro.contas_query_changed_signal.connect(self._on_contas_query_changed)
        if hasattr(self.financeiro, "export_signal"):
            self.financeiro.export_signal.connect(self.exportar_financeiro_filtrado)
        if hasattr(self.financeiro, "contas_export_signal"):
            self.financeiro.contas_export_signal.connect(self.exportar_financeiro_filtrado)
        if hasattr(self.financeiro, "contas_action_signal"):
            self.financeiro.contas_action_signal.connect(self._handle_contas_pagar_action)

        self.cadastro.salvar_signal.connect(self.salvar_cliente_no_banco)
        self.cadastro_empresa.salvar_signal.connect(self.salvar_empresa_no_banco)
        self.pagamento.registrar_signal.connect(self.registrar_pagamento_no_banco)

        if hasattr(self.cadastro, "on_check_matricula_exists"):
            self.cadastro.on_check_matricula_exists = lambda mat: db.matricula_existe(int(mat))
        if hasattr(self.cadastro, "on_find_cliente_por_cpf"):
            self.cadastro.on_find_cliente_por_cpf = lambda cpf: db.buscar_cliente_por_cpf(cpf)

        self.pagamento.on_preview_request = self.pagamento_preview_por_cpf
        self.pagamento.on_check_duplicate = self.pagamento_existe_por_cliente_mes
        self.pagamento.on_search_name_request = self.buscar_clientes_por_nome
        self.pagamento.on_preview_empresa_request = self.pagamento_preview_por_cnpj
        self.pagamento.on_check_duplicate_empresa = self.pagamento_existe_por_empresa_mes
        self.pagamento.on_search_empresa_name_request = self.buscar_empresas_por_nome

        if hasattr(self.listar, "editar_signal"):
            self.listar.editar_signal.connect(self.editar_cliente_por_mat)
        if hasattr(self.listar, "excluir_signal"):
            self.listar.excluir_signal.connect(self.excluir_cliente_por_mat)
        if hasattr(self.listar, "cancelar_plano_signal"):
            self.listar.cancelar_plano_signal.connect(self.cancelar_plano_cliente_por_mat)
        if hasattr(self.listar, "renovar_contrato_signal"):
            self.listar.renovar_contrato_signal.connect(self.renovar_contrato_cliente_por_mat)
        if hasattr(self.listar, "renovar_contratos_signal"):
            self.listar.renovar_contratos_signal.connect(self.renovar_contratos_marcados)
        if hasattr(self.listar, "reajuste_planos_signal"):
            self.listar.reajuste_planos_signal.connect(self.aplicar_reajuste_planos)
        if hasattr(self.listar, "enviar_email_signal"):
            self.listar.enviar_email_signal.connect(self.enviar_email_cliente)
        if hasattr(self.listar, "baixar_contrato_signal"):
            self.listar.baixar_contrato_signal.connect(self.baixar_contrato_cliente_por_mat)
        if hasattr(self.cadastro, "baixar_contrato_signal"):
            self.cadastro.baixar_contrato_signal.connect(self.baixar_contrato_cliente_por_mat)

        if hasattr(self.listar_empresas, "novo_signal"):
            self.listar_empresas.novo_signal.connect(self.ir_para_cadastro_empresa_create)
        if hasattr(self.listar_empresas, "editar_signal"):
            self.listar_empresas.editar_signal.connect(self.editar_empresa_por_id)
        if hasattr(self.listar_empresas, "excluir_signal"):
            self.listar_empresas.excluir_signal.connect(self.excluir_empresa_por_id)
        if hasattr(self.listar_empresas, "importar_signal"):
            self.listar_empresas.importar_signal.connect(self.importar_empresas_planilha)

    def buscar_clientes_por_nome(self, nome: str) -> list[dict]:
        rows = db.buscar_clientes_por_nome(nome, limit=20)
        return [{"id": r[0], "nome": r[1], "cpf": r[2]} for r in rows]

    def buscar_empresas_por_nome(self, nome: str) -> list[dict]:
        rows = db.buscar_empresas_por_nome(nome, limit=20)
        return [{"id": r[0], "nome": r[1], "cnpj": r[2]} for r in rows]

    def _current_user_for_preferences(self) -> str:
        username = str(self._usuario_atual or "").strip()
        if not username and hasattr(self.login, "username_input"):
            try:
                username = str(self.login.username_input.text() or "").strip()
            except Exception:
                username = ""
        if username:
            return username
        role = str(self._nivel_usuario or "").strip().lower()
        return f"role:{role or 'anon'}"

    def _apply_layout_density(self, density: str):
        mode = str(density or "").strip().lower()
        if mode not in {"normal", "compact"}:
            mode = "normal"
        self._layout_density = mode

        try:
            if self._performance_mode:
                margins = (0, 0, 0, 0)
            elif mode == "compact":
                margins = (10, 10, 10, 10)
            else:
                margins = (18, 18, 18, 18)
            if hasattr(self, "_root_layout"):
                self._root_layout.setContentsMargins(*margins)
                self._root_layout.setSpacing(0)
        except Exception:
            logger.debug("Falha ao aplicar densidade no layout raiz.", exc_info=True)

        for view, method in (
            (self.dashboard, "set_density"),
            (self.financeiro, "set_density"),
        ):
            if hasattr(view, method):
                try:
                    getattr(view, method)(mode)
                except Exception:
                    logger.debug("Falha ao aplicar densidade na view %s.", str(type(view).__name__), exc_info=True)

    def _apply_user_preferences(self, force_load: bool = False):
        _ = force_load

        # Preferências removidas: valores globais fixos.
        refresh_s = 60
        try:
            self._auto_refresh_timer.setInterval(int(refresh_s * 1000))
        except Exception:
            logger.debug("Falha ao aplicar intervalo de auto refresh.", exc_info=True)

        period_key = "month"
        self._dashboard_period = period_key

        self._apply_layout_density("normal")

        apply_period = True
        if apply_period and hasattr(self.dashboard, "set_period"):
            try:
                self.dashboard.set_period(period_key, emit_signal=False)
            except Exception:
                logger.debug("Falha ao aplicar período padrão do dashboard.", exc_info=True)

        # Tamanho padrão das grades financeiras.
        fin_page_size = 50
        contas_page_size = 50
        if hasattr(self.financeiro, "set_page_sizes"):
            try:
                self.financeiro.set_page_sizes(fin_page_size, contas_page_size)
            except Exception:
                logger.debug("Falha ao aplicar page size no financeiro.", exc_info=True)

    def _on_finance_query_changed(self, query: dict):
        safe_query = self._normalize_finance_query(query)
        self._queue_save_finance_preferences(finance_query=safe_query)
        self.atualizar_financeiro_async(
            self.financeiro.current_month(),
            force=False,
            query=safe_query,
        )

    def _on_contas_query_changed(self, query: dict):
        safe_query = dict(query or {})
        self._queue_save_finance_preferences(contas_query=safe_query)
        self.atualizar_contas_pagar_async(
            self.financeiro.current_month(),
            force=False,
            query=safe_query,
        )

    def _queue_save_finance_preferences(
        self,
        *,
        finance_query: dict | None = None,
        contas_query: dict | None = None,
    ):
        if finance_query is not None:
            self._finance_pref_query = self._normalize_finance_query(finance_query)
        if contas_query is not None:
            self._contas_pref_query = dict(contas_query or {})
        self._finance_prefs_save_timer.start()

    def _persist_finance_preferences(self):
        user_key = self._current_user_for_preferences()
        payload = {
            "financeiro_query": dict(self._finance_pref_query or {}),
            "contas_query": dict(self._contas_pref_query or {}),
        }
        try:
            db.salvar_preferencias_financeiro_usuario(user_key, payload)
        except Exception:
            logger.debug("Falha ao salvar preferências de filtros do financeiro.", exc_info=True)

    def _apply_finance_preferences_to_view(self):
        user_key = self._current_user_for_preferences()
        if self._finance_prefs_loaded_user == user_key:
            return

        try:
            prefs = db.obter_preferencias_financeiro_usuario(user_key) or {}
        except Exception:
            logger.debug("Falha ao carregar preferências de filtros do financeiro.", exc_info=True)
            prefs = {}

        fin_q = dict(prefs.get("financeiro_query") or {})
        contas_q = dict(prefs.get("contas_query") or {})
        if not fin_q and hasattr(self.financeiro, "current_query"):
            fin_q = self._normalize_finance_query(self.financeiro.current_query())
        else:
            fin_q = self._normalize_finance_query(fin_q)
        if not contas_q and hasattr(self.financeiro, "current_contas_query"):
            contas_q = dict(self.financeiro.current_contas_query() or {})

        self._finance_pref_query = dict(fin_q)
        self._contas_pref_query = dict(contas_q)

        try:
            if hasattr(self.financeiro, "apply_saved_query"):
                self.financeiro.apply_saved_query(fin_q, emit_remote=False)
            if hasattr(self.financeiro, "apply_saved_contas_query"):
                self.financeiro.apply_saved_contas_query(contas_q, emit_remote=False)
        except Exception:
            logger.debug("Falha ao aplicar preferências de filtros no financeiro.", exc_info=True)

        self._finance_prefs_loaded_user = user_key

    # ============================
    # Rotas
    # ============================
    @staticmethod
    def _month_options(count: int = 12) -> list[str]:
        now = datetime.now()
        y, m = now.year, now.month
        out: list[str] = []
        for _ in range(max(1, int(count))):
            out.append(f"{y:04d}-{m:02d}")
            m -= 1
            if m <= 0:
                m = 12
                y -= 1
        return out

    def ir_para_dashboard(self, nivel: str):
        self._nivel_usuario = str(nivel or "")
        self._usuario_atual = self._current_user_for_preferences()
        self._finance_prefs_loaded_user = ""
        self._shortcut_backup.setEnabled(self._can_backup())
        if hasattr(self.dashboard, "set_nivel_usuario"):
            self.dashboard.set_nivel_usuario(nivel)
        if hasattr(self.financeiro, "set_nivel_usuario"):
            self.financeiro.set_nivel_usuario(nivel)
        if hasattr(self.listar, "set_nivel_usuario"):
            self.listar.set_nivel_usuario(nivel)
        if hasattr(self.listar_empresas, "set_nivel_usuario"):
            self.listar_empresas.set_nivel_usuario(nivel)
        self._apply_user_preferences(force_load=True)

        self.stack.setCurrentWidget(self.dashboard)
        self._run_status_sync_async(force=True, reason="login")
        self.atualizar_dashboard_async()

    def ir_para_login(self):
        self._nivel_usuario = ""
        self._usuario_atual = ""
        self._finance_prefs_loaded_user = ""
        self._finance_pref_query = {}
        self._contas_pref_query = {}
        self._shortcut_backup.setEnabled(False)
        if hasattr(self.dashboard, "set_nivel_usuario"):
            self.dashboard.set_nivel_usuario("")
        if hasattr(self.financeiro, "set_nivel_usuario"):
            self.financeiro.set_nivel_usuario("")
        if hasattr(self.listar, "set_nivel_usuario"):
            self.listar.set_nivel_usuario("")
        if hasattr(self.listar_empresas, "set_nivel_usuario"):
            self.listar_empresas.set_nivel_usuario("")
        if hasattr(self.login, "prepare_for_show"):
            self.login.prepare_for_show()
        if hasattr(self.login, "username_input"):
            self.login.username_input.clear()
        if hasattr(self.login, "password_input"):
            self.login.password_input.clear()
        self.stack.setCurrentWidget(self.login)

    def ir_para_cadastro_create(self):
        if self._role() == ROLE_RECEPCAO:
            self.stack.setCurrentWidget(self.dashboard)
            self._notify_access_denied(
                "Perfil de recepção não pode cadastrar clientes.",
                popup=True,
            )
            return
        if hasattr(self.cadastro, "set_create_mode"):
            self.cadastro.set_create_mode()
        if hasattr(self.cadastro, "set_dependentes_lista"):
            self.cadastro.set_dependentes_lista([])
        self.stack.setCurrentWidget(self.cadastro)

    def ir_para_novo_contrato(self):
        self.ir_para_cadastro_create()
        if self.stack.currentWidget() is not self.cadastro:
            return
        try:
            if hasattr(self.cadastro, "nome") and isinstance(self.cadastro.nome, dict):
                self.cadastro.nome["input"].setFocus()
        except Exception:
            pass

    def ir_para_cadastro_empresa_create(self):
        if self._role() == ROLE_RECEPCAO:
            self.stack.setCurrentWidget(self.dashboard)
            self._notify_access_denied(
                "Perfil de recepção não pode cadastrar empresas.",
                popup=True,
            )
            return
        if hasattr(self.cadastro_empresa, "set_create_mode"):
            self.cadastro_empresa.set_create_mode()
        self.stack.setCurrentWidget(self.cadastro_empresa)

    def ir_para_pagamento(self):
        if self._role() == ROLE_RECEPCAO:
            self.stack.setCurrentWidget(self.dashboard)
            self._notify_access_denied(
                "Perfil de recepção não pode registrar pagamentos.",
                popup=True,
            )
            return
        if hasattr(self.pagamento, "set_defaults"):
            self.pagamento.set_defaults()
        self.stack.setCurrentWidget(self.pagamento)

    def ir_para_financeiro(self):
        if not self._can_access_financeiro():
            self.stack.setCurrentWidget(self.dashboard)
            self._notify_access_denied(
                "Seu perfil não tem acesso ao painel financeiro.",
                popup=True,
            )
            return

        opts = self._month_options(14)
        current = datetime.now().strftime("%Y-%m")
        if hasattr(self.financeiro, "set_month_options"):
            self.financeiro.set_month_options(opts, current)
        if hasattr(self.financeiro, "set_contas_month_options"):
            self.financeiro.set_contas_month_options(opts, current)
        if hasattr(self.financeiro, "set_nivel_usuario"):
            self.financeiro.set_nivel_usuario(self._nivel_usuario)
        self._apply_finance_preferences_to_view()

        self.stack.setCurrentWidget(self.financeiro)
        if hasattr(self.financeiro, "_emit_refresh"):
            self.financeiro._emit_refresh()

    def ir_para_listar(self):
        # Abre listagem padrão sempre na primeira página, sem filtros presos
        # da sessão anterior (evita cenário de exibir 1 único cliente).
        if hasattr(self.listar, "open_with_filters"):
            self.listar.open_with_filters(search_text="", status="", pagamento="", page=0)
        elif hasattr(self.listar, "reload"):
            self.listar.reload()
        self.stack.setCurrentWidget(self.listar)

    def ir_para_listar_empresas(self):
        if hasattr(self.listar_empresas, "reload"):
            self.listar_empresas.reload()
        self.stack.setCurrentWidget(self.listar_empresas)

    def ir_para_relatorios(self):
        reports_root = db.get_app_data_dir() / "reports"
        if hasattr(self.relatorios, "set_reports_root"):
            self.relatorios.set_reports_root(reports_root)
        if hasattr(self.relatorios, "reload"):
            self.relatorios.reload()
        self.stack.setCurrentWidget(self.relatorios)

    def ir_para_listar_filtrado(self, search_text: str = "", status: str = "", pagamento: str = ""):
        if hasattr(self.listar, "open_with_filters"):
            self.listar.open_with_filters(
                search_text=search_text or "",
                status=status or "",
                pagamento=pagamento or "",
                page=0,
            )
        else:
            if hasattr(self.listar, "reload"):
                self.listar.reload()
        self.stack.setCurrentWidget(self.listar)

    def executar_busca_global(self, texto: str):
        query = str(texto or "").strip()
        if not query:
            self.ir_para_listar()
            return

        clientes: list[tuple] = []
        empresas: list[tuple] = []
        conn = None
        try:
            conn = db.connect()
            cur = conn.cursor()
            like = f"%{query}%"
            cur.execute(
                """
                SELECT id, nome, cpf
                FROM clientes
                WHERE CAST(id AS TEXT) ILIKE ?
                   OR nome ILIKE ?
                   OR cpf ILIKE ?
                ORDER BY nome ASC
                LIMIT 8
                """,
                (like, like, like),
            )
            clientes = list(cur.fetchall() or [])

            cur.execute(
                """
                SELECT id, nome, cnpj
                FROM empresas
                WHERE CAST(id AS TEXT) ILIKE ?
                   OR nome ILIKE ?
                   OR cnpj ILIKE ?
                ORDER BY nome ASC
                LIMIT 8
                """,
                (like, like, like),
            )
            empresas = list(cur.fetchall() or [])
        except Exception as exc:
            msg = f"Falha na busca global: {exc}"
            if hasattr(self.dashboard, "show_error"):
                self.dashboard.show_error(msg)
            return
        finally:
            try:
                if conn:
                    conn.close()
            except Exception:
                pass

        if not clientes and not empresas:
            if hasattr(self.dashboard, "show_error"):
                self.dashboard.show_error("Nenhum resultado encontrado na busca global.")
            return

        digits = "".join(ch for ch in query if ch.isdigit())
        prioriza_empresa = len(digits) >= 14
        prioriza_cliente = len(digits) == 11

        if (empresas and not clientes) or (prioriza_empresa and empresas):
            if hasattr(self.listar_empresas, "open_with_filters"):
                self.listar_empresas.open_with_filters(search_text=query, forma_pagamento="", status_pagamento="")
            elif hasattr(self.listar_empresas, "reload"):
                self.listar_empresas.reload()
            self.stack.setCurrentWidget(self.listar_empresas)
            if hasattr(self.listar_empresas, "_show_message"):
                self.listar_empresas._show_message(
                    f"Busca global: {len(empresas)} empresa(s) encontrada(s).",
                    ok=True,
                    ms=2200,
                )
            return

        if (clientes and not empresas) or (prioriza_cliente and clientes):
            self.ir_para_listar_filtrado(search_text=query, status="", pagamento="")
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message(
                    f"Busca global: {len(clientes)} cliente(s) encontrado(s).",
                    ok=True,
                    ms=2200,
                )
            return

        # Se houver resultados mistos, abre clientes por padrão e informa os totais.
        self.ir_para_listar_filtrado(search_text=query, status="", pagamento="")
        if hasattr(self.listar, "_show_message"):
            self.listar._show_message(
                f"Busca global: {len(clientes)} cliente(s) e {len(empresas)} empresa(s). "
                "Use a tela de Empresas para ver os resultados empresariais.",
                ok=True,
                ms=3200,
            )

    # ============================
    # Auto refresh
    # ============================
    def _auto_refresh_tick(self):
        if self.stack.currentWidget() is self.dashboard:
            self.atualizar_dashboard_async()
        elif self.stack.currentWidget() is self.financeiro and self._can_access_financeiro():
            self.atualizar_financeiro_async(
                self.financeiro.current_month(),
                query=(self.financeiro.current_query() if hasattr(self.financeiro, "current_query") else None),
            )
        elif self.stack.currentWidget() is self.listar:
            if hasattr(self.listar, "reload"):
                self.listar.reload()
        elif self.stack.currentWidget() is self.listar_empresas:
            if hasattr(self.listar_empresas, "reload"):
                self.listar_empresas.reload()
        elif self.stack.currentWidget() is self.financeiro:
            self.stack.setCurrentWidget(self.dashboard)

    def _on_global_refresh(self):
        if self.stack.currentWidget() is self.financeiro and self._can_access_financeiro():
            self.atualizar_financeiro_async(
                self.financeiro.current_month(),
                force=True,
                query=(self.financeiro.current_query() if hasattr(self.financeiro, "current_query") else None),
            )
        elif self.stack.currentWidget() is self.listar:
            if hasattr(self.listar, "reload"):
                self.listar.reload()
        elif self.stack.currentWidget() is self.listar_empresas:
            if hasattr(self.listar_empresas, "reload"):
                self.listar_empresas.reload()
        else:
            self.atualizar_dashboard_async(force=True)

    def _ms_until_next_status_sync(self) -> int:
        now = datetime.now()
        next_run = (now + timedelta(days=1)).replace(hour=0, minute=0, second=5, microsecond=0)
        ms = int((next_run - now).total_seconds() * 1000)
        return max(30_000, ms)

    def _schedule_next_status_sync(self):
        try:
            self._status_sync_timer.stop()
            self._status_sync_timer.start(self._ms_until_next_status_sync())
        except Exception:
            pass

    def _on_status_sync_daily_tick(self):
        self._run_status_sync_async(force=True, reason="daily")
        self._schedule_next_status_sync()

    def _run_status_sync_async(self, *, force: bool = False, reason: str = ""):
        if self._status_sync_inflight:
            self._status_sync_pending_force = self._status_sync_pending_force or bool(force)
            return

        self._status_sync_inflight = True
        why = str(reason or "").strip().lower()
        worker = _Worker(lambda: db.sincronizar_status_pagamento_clientes(force=bool(force)))
        self._status_sync_workers.append(worker)

        def _cleanup():
            self._status_sync_inflight = False
            try:
                self._status_sync_workers.remove(worker)
            except Exception:
                pass

        def _flush_pending_if_any():
            if not self._status_sync_pending_force:
                return
            force_pending = bool(self._status_sync_pending_force)
            self._status_sync_pending_force = False
            QTimer.singleShot(0, lambda: self._run_status_sync_async(force=force_pending, reason="pending"))

        def _refresh_current_view_if_needed():
            current = self.stack.currentWidget()
            if current is self.dashboard:
                self._invalidate_dashboard_cache()
                self.atualizar_dashboard_async(force=True)
                return
            if current is self.listar and hasattr(self.listar, "reload"):
                self.listar.reload()
                return
            if current is self.financeiro and self._can_access_financeiro():
                self.atualizar_financeiro_async(
                    self.financeiro.current_month(),
                    force=True,
                    query=(self.financeiro.current_query() if hasattr(self.financeiro, "current_query") else None),
                )
                return

        def _on_result(result: dict):
            _cleanup()
            out = dict(result or {})
            updated = int(out.get("atualizados", 0) or 0)
            skipped = bool(out.get("skipped", False))
            if updated > 0:
                _refresh_current_view_if_needed()
            if updated > 0 or why in {"startup", "daily"}:
                logger.info(
                    "Sincronização de status dos clientes concluída: mes=%s atrasados=%s atualizados=%s skipped=%s reason=%s",
                    str(out.get("mes_ref") or ""),
                    int(out.get("atrasados", 0) or 0),
                    updated,
                    skipped,
                    why or "-",
                )
            _flush_pending_if_any()

        def _on_error(error_msg: str):
            _cleanup()
            logger.warning(
                "Falha ao sincronizar status de pagamento dos clientes (%s): %s",
                why or "manual",
                str(error_msg or ""),
            )
            _flush_pending_if_any()

        worker.signals.result.connect(_on_result)
        worker.signals.error.connect(_on_error)
        self._thread_pool.start(worker)

    # ============================
    # Backup
    # ============================
    def fazer_backup(self):
        if not self._can_backup():
            self._notify_access_denied(
                "Apenas administradores podem gerar backup.",
                popup=True,
            )
            return
        try:
            path = db.backup_db()
            self._record_activity(
                "Backup manual concluído",
                detail=Path(path).name if str(path or "").strip() else "",
                level="success",
                source="backup",
            )
            QMessageBox.information(
                self,
                "Backup criado",
                f"Backup salvo com sucesso em:\n{path}\n\n(Dica: Ctrl+B cria backup a qualquer momento)"
            )
        except Exception as e:
            safe_err = _sanitize_error_text(str(e))
            self._record_activity(
                "Backup manual falhou",
                detail=safe_err,
                level="warn",
                source="backup",
            )
            QMessageBox.critical(self, "Falha no backup", f"NÃ£o foi possÃ­vel criar backup.\n\nDetalhes: {safe_err}")

    def closeEvent(self, event):
        if not _env_flag("MEDCONTRACT_AUTO_BACKUP_ON_EXIT", False):
            event.accept()
            return
        try:
            db.backup_db()
            event.accept()
        except Exception as e:
            safe_err = _sanitize_error_text(str(e))
            resp = QMessageBox.question(
                self,
                "Backup falhou",
                f"Falhou ao criar backup automÃ¡tico ao fechar.\n\nDetalhes: {safe_err}\n\nDeseja fechar mesmo assim?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            event.accept() if resp == QMessageBox.Yes else event.ignore()

    def _on_dashboard_period_changed(self, period_key: str):
        key = (period_key or "").strip().lower()
        if key not in {"month", "7d", "today"}:
            key = "month"
        self._dashboard_period = key

    def _invalidate_dashboard_cache(self, period_key: str | None = None):
        # Financeiro compartilha parte das mesmas fontes (clientes/pagamentos).
        # Ao invalidar dashboard, invalidamos também cache financeiro.
        self._invalidate_finance_cache()
        self._invalidate_contas_cache()
        if period_key is None:
            self._dashboard_cache.clear()
            return
        key = (period_key or "").strip().lower()
        if not key:
            return
        self._dashboard_cache.pop(key, None)

    def _get_cached_dashboard_payload(self, period_key: str) -> dict | None:
        if self._dashboard_cache_ttl_s <= 0:
            return None
        key = (period_key or "").strip().lower()
        if key not in {"month", "7d", "today"}:
            key = "month"
        entry = self._dashboard_cache.get(key)
        if not entry:
            return None
        cached_at, payload = entry
        age = (datetime.now() - cached_at).total_seconds()
        if age > self._dashboard_cache_ttl_s:
            self._dashboard_cache.pop(key, None)
            return None
        return payload

    def _set_cached_dashboard_payload(self, period_key: str, payload: dict):
        if self._dashboard_cache_ttl_s <= 0:
            return
        key = (period_key or "").strip().lower()
        if key not in {"month", "7d", "today"}:
            key = "month"
        self._dashboard_cache[key] = (datetime.now(), payload)

    @staticmethod
    def _normalize_finance_query(query: dict | None) -> dict:
        src = dict(query or {})
        min_value = src.get("min_value")
        max_value = src.get("max_value")
        try:
            min_value = None if min_value in (None, "") else float(min_value)
        except Exception:
            min_value = None
        try:
            max_value = None if max_value in (None, "") else float(max_value)
        except Exception:
            max_value = None
        if min_value is not None and max_value is not None and min_value > max_value:
            min_value, max_value = max_value, min_value

        sort_key = str(src.get("sort_key", "data_pagamento") or "data_pagamento").strip().lower()
        if sort_key not in {
            "data_pagamento", "mat", "nome", "cpf", "status",
            "pagamento_status", "valor_pago", "mes_referencia",
        }:
            sort_key = "data_pagamento"

        sort_dir = str(src.get("sort_dir", "desc") or "desc").strip().lower()
        sort_dir = "asc" if sort_dir == "asc" else "desc"

        try:
            page = max(0, int(src.get("page", 0) or 0))
        except Exception:
            page = 0
        try:
            page_size = max(1, int(src.get("page_size", 50) or 50))
        except Exception:
            page_size = 50
        try:
            ticket_ref = float(src.get("ticket_ref", 0.0) or 0.0)
        except Exception:
            ticket_ref = 0.0

        return {
            "page": page,
            "page_size": page_size,
            "search_doc": str(src.get("search_doc", "") or "").strip(),
            "search_name": str(src.get("search_name", "") or "").strip(),
            "status_key": str(src.get("status_key", "") or "").strip().lower(),
            "min_value": min_value,
            "max_value": max_value,
            "only_atrasados": bool(src.get("only_atrasados", False)),
            "above_ticket": bool(src.get("above_ticket", False)),
            "ticket_ref": ticket_ref,
            "only_today": bool(src.get("only_today", False)),
            "sort_key": sort_key,
            "sort_dir": sort_dir,
        }

    def _finance_cache_key(self, mes_iso: str, query: dict | None = None) -> str:
        ref = str(mes_iso or "").strip()
        if len(ref) != 7 or ref[4] != "-":
            ref = datetime.now().strftime("%Y-%m")
        normalized = self._normalize_finance_query(query)
        try:
            q_raw = json.dumps(normalized, sort_keys=True, separators=(",", ":"), ensure_ascii=False)
        except Exception:
            q_raw = str(normalized)
        return f"{ref}|{q_raw}"

    def _invalidate_finance_cache(self, mes_iso: str | None = None):
        if mes_iso is None:
            self._finance_cache.clear()
            return
        ref = str(mes_iso or "").strip()
        if len(ref) != 7 or ref[4] != "-":
            return
        prefix = f"{ref}|"
        for key in list(self._finance_cache.keys()):
            if key.startswith(prefix):
                self._finance_cache.pop(key, None)

    def _invalidate_contas_cache(self, mes_iso: str | None = None):
        if mes_iso is None:
            self._contas_cache.clear()
            return
        ref = str(mes_iso or "").strip()
        if len(ref) != 7 or ref[4] != "-":
            return
        prefix = f"{ref}|"
        for key in list(self._contas_cache.keys()):
            if key.startswith(prefix):
                self._contas_cache.pop(key, None)

    def _get_cached_financeiro_payload(self, mes_iso: str, query: dict | None = None) -> dict | None:
        if self._finance_cache_ttl_s <= 0:
            return None
        key = self._finance_cache_key(mes_iso, query=query)
        entry = self._finance_cache.get(key)
        if not entry:
            return None
        cached_at, payload = entry
        age = (datetime.now() - cached_at).total_seconds()
        if age > self._finance_cache_ttl_s:
            self._finance_cache.pop(key, None)
            return None
        return payload

    def _set_cached_financeiro_payload(self, mes_iso: str, payload: dict, query: dict | None = None):
        if self._finance_cache_ttl_s <= 0:
            return
        key = self._finance_cache_key(mes_iso, query=query)
        self._finance_cache[key] = (datetime.now(), payload)

    def _contas_cache_key(self, mes_iso: str, query: dict | None = None) -> str:
        ref = str(mes_iso or "").strip()
        if len(ref) != 7 or ref[4] != "-":
            ref = datetime.now().strftime("%Y-%m")
        query_safe = dict(query or {})
        try:
            q_raw = json.dumps(query_safe, sort_keys=True, separators=(",", ":"), ensure_ascii=False)
        except Exception:
            q_raw = str(query_safe)
        return f"{ref}|{q_raw}"

    def _get_cached_contas_payload(self, mes_iso: str, query: dict | None = None) -> dict | None:
        if self._finance_cache_ttl_s <= 0:
            return None
        key = self._contas_cache_key(mes_iso, query=query)
        entry = self._contas_cache.get(key)
        if not entry:
            return None
        cached_at, payload = entry
        age = (datetime.now() - cached_at).total_seconds()
        if age > self._finance_cache_ttl_s:
            self._contas_cache.pop(key, None)
            return None
        return payload

    def _set_cached_contas_payload(self, mes_iso: str, payload: dict, query: dict | None = None):
        if self._finance_cache_ttl_s <= 0:
            return
        key = self._contas_cache_key(mes_iso, query=query)
        self._contas_cache[key] = (datetime.now(), payload)

    def _start_tracked_worker(
        self,
        fn,
        *args,
        bucket: list,
        on_result=None,
        on_error=None,
        on_finish=None,
    ):
        worker = _Worker(fn, *args)
        bucket.append(worker)

        def _cleanup():
            try:
                bucket.remove(worker)
            except Exception:
                pass
            if callable(on_finish):
                try:
                    on_finish()
                except Exception:
                    pass

        def _handle_result(payload):
            _cleanup()
            if callable(on_result):
                on_result(payload)

        def _handle_error(error_msg: str):
            _cleanup()
            if callable(on_error):
                on_error(error_msg)

        worker.signals.result.connect(_handle_result)
        worker.signals.error.connect(_handle_error)
        self._thread_pool.start(worker)
        return worker

    # ============================
    # Dashboard refresh
    # ============================
    def atualizar_dashboard_async(self, force: bool = False):
        if self._refresh_inflight:
            # Coalesce: mantém no máximo 1 atualização pendente enquanto outra roda.
            self._dashboard_refresh_pending = True
            self._dashboard_refresh_pending_force = (
                self._dashboard_refresh_pending_force or bool(force)
            )
            return

        period_key = (self._dashboard_period or "month").strip().lower()
        if period_key not in {"month", "7d", "today"}:
            period_key = "month"

        if not force:
            cached = self._get_cached_dashboard_payload(period_key)
            if cached is not None:
                self._apply_dashboard_payload(cached)
                return

        self._refresh_inflight = True

        if hasattr(self.dashboard, "set_refresh_state"):
            try:
                self.dashboard.set_refresh_state(True)
            except Exception:
                pass

        w = _Worker(self._compute_dashboard_payload, period_key)
        w.signals.result.connect(self._apply_dashboard_payload)
        w.signals.error.connect(self._on_dashboard_refresh_error)
        self._thread_pool.start(w)

    def _schedule_dashboard_refresh_if_pending(self):
        if not self._dashboard_refresh_pending:
            return
        force = bool(self._dashboard_refresh_pending_force)
        self._dashboard_refresh_pending = False
        self._dashboard_refresh_pending_force = False
        QTimer.singleShot(0, lambda: self.atualizar_dashboard_async(force=force))

    def _on_dashboard_refresh_error(self, _msg: str):
        self._refresh_inflight = False
        msg = (_msg or "").strip() or "Não foi possível atualizar o dashboard."
        if hasattr(self.dashboard, "set_refresh_state"):
            try:
                self.dashboard.set_refresh_state(False, "Falha na atualização")
                self.dashboard.show_error(msg)
            except Exception:
                pass
        self._schedule_dashboard_refresh_if_pending()

    def _compute_dashboard_payload(self, period: str = "month") -> dict:
        payload = dashboard_payload_service.compute_dashboard_payload(
            db,
            period,
            iso_to_mes_ref_br_fn=iso_to_mes_ref_br,
            log_debug=lambda message: logger.debug(str(message or ""), exc_info=True),
            alert_user=self._current_user_for_preferences(),
        )
        try:
            resumo = dict((payload or {}).get("resumo", {}) or {})
            if self._export_history:
                resumo["ultima_export"] = self._export_history[0].get("when", "-")
            payload["resumo"] = resumo
        except Exception:
            logger.debug("Falha ao enriquecer payload do dashboard com dados locais.", exc_info=True)

        payload["export_history"] = list(self._export_history)
        payload["recent_activities"] = list(self._activity_history[:6])
        return payload

    def _build_operational_summary_text(self, payload: dict) -> str:
        return _build_operational_summary_payload(payload, now=datetime.now())

    def _build_due_digest_text(self, payload: dict) -> str:
        now = datetime.now()
        generated_at = now.strftime("%d/%m/%Y %H:%M:%S")
        period_desc = str(payload.get("period_desc") or "-")
        sc = dict(payload.get("status_counts", {}) or {})
        forecast = payload.get("finance_forecast")
        if not isinstance(forecast, dict):
            forecast = payload.get("contratos_mes", {}) or {}
        forecast = dict(forecast or {})

        atrasados = int(sc.get("atrasados", 0) or 0)
        qtd_7d = int(forecast.get("qtd_7d", 0) or 0)
        qtd_15d = int(forecast.get("qtd_15d", 0) or 0)
        qtd_30d = int(forecast.get("qtd_30d", 0) or 0)
        entrada_30d = float(forecast.get("entrada_30d", 0.0) or 0.0)
        risco_30d = float(forecast.get("risco_30d", 0.0) or 0.0)
        liquido_30d = float(forecast.get("previsao_liquida_30d", 0.0) or 0.0)
        risco_nivel = str(forecast.get("risco_nivel", "baixo") or "baixo").strip().lower()
        risco_pct = float(forecast.get("taxa_inadimplencia", 0.0) or 0.0) * 100.0

        nivel_legivel = {
            "baixo": "Baixo",
            "medio": "Médio",
            "alto": "Alto",
            "critico": "Crítico",
        }.get(risco_nivel, risco_nivel.title() or "Baixo")

        return (
            "MEDCONTRACT - LEMBRETE DE VENCIMENTOS E RISCO\n"
            f"Gerado em: {generated_at}\n"
            f"Período do dashboard: {period_desc}\n\n"
            "PONTOS DE ATENÇÃO\n"
            f"- Clientes em atraso: {atrasados}\n"
            f"- Vencimentos em 7 dias: {qtd_7d}\n"
            f"- Vencimentos em 15 dias: {qtd_15d}\n"
            f"- Vencimentos em 30 dias: {qtd_30d}\n\n"
            "PROJEÇÃO FINANCEIRA (30 DIAS)\n"
            f"- Entrada prevista: {entrada_30d:.2f}\n"
            f"- Risco projetado: {risco_30d:.2f}\n"
            f"- Líquido previsto: {liquido_30d:.2f}\n"
            f"- Taxa estimada de inadimplência: {risco_pct:.1f}% ({nivel_legivel})\n"
        )

    def _maybe_generate_operational_summary(self, payload: dict):
        day_key = datetime.now().strftime("%Y-%m-%d")
        if self._last_operational_summary_date == day_key:
            return
        try:
            reports_dir = db.get_app_data_dir() / "reports"
            reports_dir.mkdir(parents=True, exist_ok=True)
            dst = reports_dir / f"resumo_operacional_{day_key}.txt"
            dst.write_text(self._build_operational_summary_text(payload), encoding="utf-8")
            self._last_operational_summary_date = day_key
            self._record_activity(
                "Resumo operacional gerado",
                detail=dst.name,
                level="success",
                source="resumo",
            )
        except Exception as exc:
            logger.warning("Falha ao gerar resumo operacional: %s", exc)

    def _maybe_generate_due_digest(self, payload: dict):
        day_key = datetime.now().strftime("%Y-%m-%d")
        if self._last_due_digest_date == day_key:
            return
        try:
            reports_dir = db.get_app_data_dir() / "reports"
            reports_dir.mkdir(parents=True, exist_ok=True)
            dst = reports_dir / f"lembrete_vencimentos_{day_key}.txt"
            dst.write_text(self._build_due_digest_text(payload), encoding="utf-8")
            self._last_due_digest_date = day_key
            self._record_activity(
                "Lembrete diário de vencimentos gerado",
                detail=dst.name,
                level="success",
                source="automacao",
            )
        except Exception as exc:
            logger.warning("Falha ao gerar lembrete diário de vencimentos: %s", exc)

    def _build_jobs_status(self, resumo: dict | None = None) -> dict:
        now = datetime.now()
        try:
            run_hour = int((os.getenv("MEDCONTRACT_AUTO_EXPORT_HOUR") or "8").strip())
        except Exception:
            run_hour = 8
        return _build_jobs_status_payload(
            now=now,
            resumo=resumo or {},
            backup_dir=db.get_backup_dir(),
            reports_dir=db.get_app_data_dir() / "reports",
            export_history=list(self._export_history),
            last_auto_export_key=str(self._last_auto_export_key or ""),
            last_operational_summary_date=str(self._last_operational_summary_date or ""),
            auto_export_enabled=_env_flag("MEDCONTRACT_AUTO_EXPORT_ENABLED", False),
            auto_export_hour=run_hour,
        )

    def _refresh_dashboard_operational_alerts(self, payload: dict, jobs_status: dict):
        if not hasattr(self.dashboard, "clear_alerts") or not hasattr(self.dashboard, "add_alert"):
            return

        try:
            self.dashboard.clear_alerts()
        except Exception:
            return

        sc = dict(payload.get("status_counts", {}) or {})
        forecast = payload.get("finance_forecast")
        if not isinstance(forecast, dict):
            forecast = payload.get("contratos_mes", {}) or {}
        forecast = dict(forecast or {})
        jobs = dict(jobs_status or {})

        try:
            atrasados = int(sc.get("atrasados", 0) or 0)
        except Exception:
            atrasados = 0
        if atrasados > 0:
            self.dashboard.add_alert("warn", f"Clientes em atraso: {atrasados}. Priorize cobrança/regularização.")

        try:
            venc_7d = int(forecast.get("qtd_7d", 0) or 0)
        except Exception:
            venc_7d = 0
        if venc_7d > 0:
            self.dashboard.add_alert("info", f"Próximos vencimentos em 7 dias: {venc_7d}.")

        try:
            taxa_risco = float(forecast.get("taxa_inadimplencia", 0.0) or 0.0)
        except Exception:
            taxa_risco = 0.0
        if taxa_risco >= 0.2:
            self.dashboard.add_alert("danger", f"Risco de inadimplência alto ({taxa_risco * 100:.1f}%).")
        elif taxa_risco >= 0.12:
            self.dashboard.add_alert("warn", f"Risco de inadimplência moderado ({taxa_risco * 100:.1f}%).")

        auto_job = jobs.get("autoexport", {}) if isinstance(jobs, dict) else {}
        auto_level = ""
        auto_text = ""
        if isinstance(auto_job, dict):
            auto_level = str(auto_job.get("level", "") or "").strip().lower()
            auto_text = str(auto_job.get("text", "") or "").strip()
        elif isinstance(auto_job, str):
            auto_text = auto_job.strip()
        if auto_level in {"warn", "warning", "error", "failed"} and auto_text:
            self.dashboard.add_alert("warn", f"Autoexport: {auto_text}.")

    def _apply_dashboard_payload(self, payload: dict):
        self._refresh_inflight = False
        period_key = str(payload.get("period_key") or self._dashboard_period or "month")
        self._set_cached_dashboard_payload(period_key, payload)

        if hasattr(self.dashboard, "set_refresh_state"):
            try:
                period_desc = payload.get("period_desc", "")
                self.dashboard.set_refresh_state(False, str(period_desc or "Atualizado"))
            except Exception:
                pass

        sc = payload.get("status_counts", {}) or {}
        if hasattr(self.dashboard, "set_status_counts"):
            self.dashboard.set_status_counts(sc.get("ativos", 0), sc.get("atrasados", 0), sc.get("inativos", 0))

        lm = payload.get("live_metrics", {}) or {}
        if hasattr(self.dashboard, "set_live_metrics"):
            self.dashboard.set_live_metrics(lm)

        rs = payload.get("resumo", {}) or {}
        if hasattr(self.dashboard, "set_resumo_do_dia"):
            self.dashboard.set_resumo_do_dia(rs)

        series = payload.get("series", []) or []
        if hasattr(self.dashboard, "set_chart_series"):
            self.dashboard.set_chart_series(series, payload.get("period_chart_label"))

        cm = payload.get("finance_forecast")
        if not isinstance(cm, dict):
            cm = payload.get("contratos_mes", {}) or {}
        if hasattr(self.dashboard, "set_contratos_mes"):
            self.dashboard.set_contratos_mes(cm)

        if hasattr(self.dashboard, "set_export_history"):
            self.dashboard.set_export_history(payload.get("export_history", []) or [])
        if hasattr(self.dashboard, "set_recent_activities"):
            self.dashboard.set_recent_activities(payload.get("recent_activities", []) or [])

        self._maybe_generate_operational_summary(payload)
        self._maybe_generate_due_digest(payload)
        jobs_status = self._build_jobs_status(rs)
        if hasattr(self.dashboard, "set_jobs_status"):
            self.dashboard.set_jobs_status(jobs_status)
        self._refresh_dashboard_operational_alerts(payload, jobs_status)
        self._schedule_dashboard_refresh_if_pending()

    # ============================
    # Financeiro refresh
    # ============================
    def atualizar_financeiro_async(
        self,
        mes_iso: str | None = None,
        force: bool = False,
        query: dict | None = None,
    ):
        if not self._can_access_financeiro():
            return

        ref = (mes_iso or "").strip()
        if len(ref) != 7 or ref[4] != "-":
            ref = datetime.now().strftime("%Y-%m")
        q = self._normalize_finance_query(query)

        if self._finance_inflight:
            # Coalesce: guarda apenas a última solicitação pendente.
            self._finance_refresh_pending = True
            self._finance_refresh_pending_ref = ref
            self._finance_refresh_pending_force = (
                self._finance_refresh_pending_force or bool(force)
            )
            self._finance_refresh_pending_query = q
            return

        if not force:
            cached = self._get_cached_financeiro_payload(ref, query=q)
            if cached is not None:
                try:
                    self.financeiro.set_payload(cached)
                    self.financeiro.set_loading(False)
                except Exception:
                    logger.debug("Falha ao aplicar payload financeiro em cache.", exc_info=True)
                return

        self._finance_inflight = True
        try:
            self.financeiro.set_loading(True)
        except Exception:
            logger.debug("Falha ao marcar loading do financeiro.", exc_info=True)

        w = _Worker(self._compute_financeiro_payload, ref, q)
        w.signals.result.connect(self._apply_financeiro_payload)
        w.signals.error.connect(self._on_financeiro_refresh_error)
        self._thread_pool.start(w)

    def _schedule_financeiro_refresh_if_pending(self):
        if not self._finance_refresh_pending:
            return
        ref = str(self._finance_refresh_pending_ref or "").strip()
        force = bool(self._finance_refresh_pending_force)
        query = dict(self._finance_refresh_pending_query or {})
        self._finance_refresh_pending = False
        self._finance_refresh_pending_ref = ""
        self._finance_refresh_pending_force = False
        self._finance_refresh_pending_query = None
        QTimer.singleShot(0, lambda: self.atualizar_financeiro_async(ref, force=force, query=query))

    def _on_financeiro_refresh_error(self, msg: str):
        self._finance_inflight = False
        try:
            self.financeiro.set_loading(False)
            self.financeiro.show_error(msg or "Falha ao atualizar painel financeiro.")
        except Exception:
            logger.debug("Falha ao exibir erro no painel financeiro.", exc_info=True)
        self._schedule_financeiro_refresh_if_pending()

    def _compute_financeiro_payload(self, mes_iso: str, query: dict | None = None) -> dict:
        q = self._normalize_finance_query(query)
        return finance_payload_service.compute_financeiro_payload(db, mes_iso, q)

    def _apply_financeiro_payload(self, payload: dict):
        self._finance_inflight = False
        try:
            ref = str((payload or {}).get("mes_ref", "") or "")
            if ref:
                self._set_cached_financeiro_payload(
                    ref,
                    payload or {},
                    query=(payload or {}).get("query"),
                )
        except Exception:
            logger.debug("Falha ao salvar cache do financeiro.", exc_info=True)
        try:
            self.financeiro.set_payload(payload or {})
            self.financeiro.set_loading(False)
        except Exception:
            logger.debug("Falha ao aplicar payload do financeiro.", exc_info=True)
        self._schedule_financeiro_refresh_if_pending()

    # ============================
    # Contas a Pagar refresh
    # ============================
    def atualizar_contas_pagar_async(
        self,
        mes_iso: str | None = None,
        force: bool = False,
        query: dict | None = None,
    ):
        if not self._can_access_financeiro():
            return

        ref = (mes_iso or "").strip()
        if len(ref) != 7 or ref[4] != "-":
            ref = datetime.now().strftime("%Y-%m")
        q = dict(query or {})

        if self._contas_inflight:
            self._contas_refresh_pending = True
            self._contas_refresh_pending_ref = ref
            self._contas_refresh_pending_force = self._contas_refresh_pending_force or bool(force)
            self._contas_refresh_pending_query = q
            return

        if not force:
            cached = self._get_cached_contas_payload(ref, query=q)
            if cached is not None:
                try:
                    self.financeiro.set_contas_payload(cached)
                    self.financeiro.set_contas_loading(False)
                except Exception:
                    logger.debug("Falha ao aplicar payload de contas em cache.", exc_info=True)
                return

        self._contas_inflight = True
        try:
            self.financeiro.set_contas_loading(True)
        except Exception:
            logger.debug("Falha ao marcar loading de contas a pagar.", exc_info=True)

        w = _Worker(self._compute_contas_pagar_payload, ref, q)
        w.signals.result.connect(self._apply_contas_pagar_payload)
        w.signals.error.connect(self._on_contas_pagar_refresh_error)
        self._thread_pool.start(w)

    def _schedule_contas_pagar_refresh_if_pending(self):
        if not self._contas_refresh_pending:
            return
        ref = str(self._contas_refresh_pending_ref or "").strip()
        force = bool(self._contas_refresh_pending_force)
        query = dict(self._contas_refresh_pending_query or {})
        self._contas_refresh_pending = False
        self._contas_refresh_pending_ref = ""
        self._contas_refresh_pending_force = False
        self._contas_refresh_pending_query = None
        QTimer.singleShot(0, lambda: self.atualizar_contas_pagar_async(ref, force=force, query=query))

    def _on_contas_pagar_refresh_error(self, msg: str):
        self._contas_inflight = False
        try:
            self.financeiro.set_contas_loading(False)
            self.financeiro.show_contas_error(msg or "Falha ao atualizar contas a pagar.")
        except Exception:
            logger.debug("Falha ao exibir erro de contas a pagar.", exc_info=True)
        self._schedule_contas_pagar_refresh_if_pending()

    def _compute_contas_pagar_payload(self, mes_iso: str, query: dict | None = None) -> dict:
        q = dict(query or {})
        return finance_payload_service.compute_contas_pagar_payload(db, mes_iso, q)

    def _apply_contas_pagar_payload(self, payload: dict):
        self._contas_inflight = False
        try:
            ref = str((payload or {}).get("mes_ref", "") or "")
            if ref:
                self._set_cached_contas_payload(ref, payload or {}, query=(payload or {}).get("query"))
        except Exception:
            logger.debug("Falha ao salvar cache de contas a pagar.", exc_info=True)
        try:
            self.financeiro.set_contas_payload(payload or {})
            self.financeiro.set_contas_loading(False)
        except Exception:
            logger.debug("Falha ao aplicar payload de contas a pagar.", exc_info=True)
        self._schedule_contas_pagar_refresh_if_pending()

    # ============================
    # ExportaÃ§Ã£o
    # ============================
    def _choose_export_path(self, default_name: str):
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar exportaÃ§Ã£o",
            default_name,
            "Excel (*.xlsx);;CSV (*.csv)"
        )
        if not path:
            return None
        return path

    @staticmethod
    def _currency_header_indexes(headers: list[str]) -> set[int]:
        idxs: set[int] = set()
        for i, h in enumerate(headers):
            key = str(h or "").strip().lower()
            if any(k in key for k in ("valor", "receita", "total")):
                idxs.add(i)
        return idxs

    @staticmethod
    def _to_currency_number(raw):
        if raw in (None, ""):
            return None
        if isinstance(raw, (int, float)):
            return float(raw)

        s = str(raw).strip()
        if not s:
            return None
        s = s.replace("R$", "").replace("r$", "").replace(" ", "")
        s = re.sub(r"[^0-9,.\-]", "", s)
        if not s or s in {"-", ".", ","}:
            return None

        neg = s.startswith("-")
        if neg:
            s = s[1:]
        if not s:
            return None

        last_comma = s.rfind(",")
        last_dot = s.rfind(".")
        if last_comma == -1 and last_dot == -1:
            num = float(s)
            return -num if neg else num

        # Caso comum de milhar sem casas decimais: 1.000 / 10.000 / 100.000
        if last_comma == -1 and s.count(".") == 1:
            int_part_raw, frac_part_raw = s.split(".", 1)
            if int_part_raw.isdigit() and frac_part_raw.isdigit() and len(frac_part_raw) == 3:
                num = float(f"{int_part_raw}{frac_part_raw}")
                return -num if neg else num
        if last_dot == -1 and s.count(",") == 1:
            int_part_raw, frac_part_raw = s.split(",", 1)
            if int_part_raw.isdigit() and frac_part_raw.isdigit() and len(frac_part_raw) == 3:
                num = float(f"{int_part_raw}{frac_part_raw}")
                return -num if neg else num

        sep_idx = max(last_comma, last_dot)
        int_part = re.sub(r"[.,]", "", s[:sep_idx]) or "0"
        frac_raw = re.sub(r"[.,]", "", s[sep_idx + 1 :])
        frac = (frac_raw + "00")[:2] if frac_raw else "00"
        num = float(f"{int_part}.{frac}")
        return -num if neg else num

    @staticmethod
    def _to_currency_br_string(raw) -> str:
        try:
            num = MainWindow._to_currency_number(raw)
        except Exception:
            num = None
        if num is None:
            return str(raw) if raw is not None else ""
        return f"{num:.2f}".replace(".", ",")

    @staticmethod
    def _iso_date_to_br(raw_date: str) -> str:
        txt = str(raw_date or "").strip()
        if len(txt) == 10 and txt[4] == "-" and txt[7] == "-":
            try:
                y, m, d = txt.split("-")
                return f"{d}/{m}/{y}"
            except Exception:
                return txt
        return txt

    @staticmethod
    def _sanitize_spreadsheet_cell(value):
        if value is None:
            return ""
        if isinstance(value, (int, float, bool)):
            return value
        text = str(value)
        if text and text[0] in {"=", "+", "-", "@", "\t", "\r", "\n"}:
            return f"'{text}"
        return text

    @staticmethod
    def _normalize_import_header(value: str) -> str:
        txt = unicodedata.normalize("NFKD", str(value or "").strip().lower())
        txt = "".join(ch for ch in txt if not unicodedata.combining(ch))
        return re.sub(r"[^a-z0-9]+", "", txt)

    def _map_empresa_import_headers(self, headers: list[str]) -> dict[int, str]:
        aliases = {
            "cnpj": "cnpj",
            "razaosocial": "nome",
            "nome": "nome",
            "empresa": "nome",
            "telefone": "telefone",
            "email": "email",
            "logradouro": "logradouro",
            "endereco": "logradouro",
            "rua": "logradouro",
            "numero": "numero",
            "bairro": "bairro",
            "cep": "cep",
            "cidade": "cidade",
            "municipio": "cidade",
            "estado": "estado",
            "uf": "estado",
            "formapagamento": "forma_pagamento",
            "pagamento": "forma_pagamento",
            "statuspagamento": "status_pagamento",
            "status": "status_pagamento",
            "diavencimento": "dia_vencimento",
            "vencimento": "dia_vencimento",
            "vencimentodia": "dia_vencimento",
            "valormensal": "valor_mensal",
            "valor": "valor_mensal",
            "mensalidade": "valor_mensal",
        }
        required = {
            "cnpj",
            "nome",
            "telefone",
            "email",
            "logradouro",
            "numero",
            "bairro",
            "cep",
            "cidade",
            "estado",
            "forma_pagamento",
            "status_pagamento",
            "dia_vencimento",
            "valor_mensal",
        }

        out: dict[int, str] = {}
        present: set[str] = set()
        for idx, header in enumerate(headers):
            token = self._normalize_import_header(header)
            field = aliases.get(token)
            if not field:
                continue
            out[idx] = field
            present.add(field)

        missing = sorted(required - present)
        if missing:
            raise RuntimeError(
                "Cabeçalhos ausentes na planilha de empresas: "
                + ", ".join(missing)
            )
        return out

    @staticmethod
    def _spreadsheet_cell_to_text(value) -> str:
        if value is None:
            return ""
        if isinstance(value, bool):
            return "1" if value else "0"
        if isinstance(value, int):
            return str(value)
        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            return str(value).replace(".", ",")
        return str(value).strip()

    def _read_empresas_planilha(self, path: str) -> list[tuple[int, dict]]:
        src = Path(path)
        ext = src.suffix.lower()
        rows: list[tuple[int, dict]] = []

        if ext == ".xlsx":
            wb = load_workbook(str(src), read_only=True, data_only=True)
            try:
                ws = wb.active
                iterator = ws.iter_rows(values_only=True)
                header_row = next(iterator, None)
                if not header_row:
                    raise RuntimeError("Planilha vazia.")
                headers = [self._spreadsheet_cell_to_text(h) for h in list(header_row)]
                mapping = self._map_empresa_import_headers(headers)
                line_no = 1
                for raw in iterator:
                    line_no += 1
                    values = list(raw or [])
                    if not values or all(not str(v or "").strip() for v in values):
                        continue
                    payload: dict[str, str] = {}
                    for idx, field in mapping.items():
                        payload[field] = self._spreadsheet_cell_to_text(values[idx] if idx < len(values) else "")
                    rows.append((line_no, payload))
            finally:
                wb.close()
            return rows

        if ext == ".csv":
            with open(src, "r", newline="", encoding="utf-8-sig") as fp:
                sample = fp.read(4096)
                fp.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=";,")
                except Exception:
                    class _FallbackDialect(csv.excel):
                        delimiter = ";"
                    dialect = _FallbackDialect()
                reader = csv.reader(fp, dialect)
                header_row = next(reader, None)
                if not header_row:
                    raise RuntimeError("Planilha CSV vazia.")
                headers = [self._spreadsheet_cell_to_text(h) for h in list(header_row)]
                mapping = self._map_empresa_import_headers(headers)
                line_no = 1
                for raw in reader:
                    line_no += 1
                    values = list(raw or [])
                    if not values or all(not str(v or "").strip() for v in values):
                        continue
                    payload: dict[str, str] = {}
                    for idx, field in mapping.items():
                        payload[field] = self._spreadsheet_cell_to_text(values[idx] if idx < len(values) else "")
                    rows.append((line_no, payload))
            return rows

        raise RuntimeError("Formato inválido. Use .xlsx ou .csv.")

    def importar_empresas_planilha(self):
        if self._role() == ROLE_RECEPCAO:
            self._notify_access_denied(
                "Perfil de recepção não pode importar empresas.",
                popup=True,
            )
            return

        path, _ = QFileDialog.getOpenFileName(
            self,
            "Importar empresas por planilha",
            "",
            "Planilhas (*.xlsx *.csv);;Excel (*.xlsx);;CSV (*.csv)",
        )
        if not path:
            return

        try:
            rows = self._read_empresas_planilha(path)
        except Exception as exc:
            QMessageBox.critical(
                self,
                "Falha na importação",
                f"Não foi possível ler a planilha.\n\nDetalhes: {exc}",
            )
            return

        if not rows:
            QMessageBox.information(
                self,
                "Importação de empresas",
                "Nenhuma linha válida foi encontrada na planilha.",
            )
            return

        confirm = QMessageBox.question(
            self,
            "Confirmar importação",
            f"Foram encontradas {len(rows)} linha(s) para importar.\n\nDeseja continuar?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if confirm != QMessageBox.Yes:
            return

        ok_count = 0
        fail_msgs: list[str] = []
        for line_no, data in rows:
            payload = dict(data or {})
            payload["modo"] = "create"
            ok, msg = empresa_controller.salvar_empresa(payload)
            if ok:
                ok_count += 1
            else:
                fail_msgs.append(f"Linha {line_no}: {msg}")

        fail_count = len(fail_msgs)
        summary = f"Importação concluída.\n\nSucesso: {ok_count}\nFalhas: {fail_count}"
        if fail_msgs:
            summary += "\n\nPrimeiras falhas:\n- " + "\n- ".join(fail_msgs[:8])
        QMessageBox.information(self, "Importação de empresas", summary)

        if ok_count > 0:
            self._invalidate_dashboard_cache()
            self.atualizar_dashboard_async()
            if hasattr(self.listar_empresas, "_show_message"):
                self.listar_empresas._show_message(
                    f"Importação finalizada: {ok_count} empresa(s) cadastrada(s).",
                    ok=True,
                    ms=2800,
                )
            if hasattr(self.listar_empresas, "reload"):
                self.listar_empresas.reload()

    def _write_csv(self, path: str, headers: list[str], rows: list[list]):
        money_cols = self._currency_header_indexes(headers)
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, delimiter=";")
            w.writerow(headers)
            for r in rows:
                out = []
                for i, value in enumerate(r):
                    if i in money_cols:
                        rendered = self._to_currency_br_string(value)
                    else:
                        rendered = value
                    out.append(self._sanitize_spreadsheet_cell(rendered))
                w.writerow(out)

    def _write_xlsx(self, path: str, headers: list[str], rows: list[list]):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter
        except Exception:
            raise RuntimeError(
                "openpyxl nÃ£o estÃ¡ disponÃ­vel. Ative o ambiente virtual correto e instale com:\n"
                r"venv\Scripts\python.exe -m pip install openpyxl"
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "RelatÃ³rio"

        # =========================
        # ESTILOS
        # =========================
        cor_titulo = "2B6C7E"
        cor_subtitulo = "64748B"
        cor_header_bg = "2B6C7E"
        cor_header_fg = "FFFFFF"
        cor_borda = "D7DEE7"
        cor_linha_par = "F8FAFC"
        cor_linha_impar = "FFFFFF"

        thin = Side(style="thin", color=cor_borda)
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

        font_title = Font(name="Segoe UI", size=16, bold=True, color=cor_titulo)
        font_subtitle = Font(name="Segoe UI", size=10, italic=False, color=cor_subtitulo)
        font_header = Font(name="Segoe UI", size=11, bold=True, color=cor_header_fg)
        font_body = Font(name="Segoe UI", size=10, color="0F172A")
        font_total = Font(name="Segoe UI", size=10, bold=True, color="0F172A")

        fill_header = PatternFill("solid", fgColor=cor_header_bg)
        fill_total = PatternFill("solid", fgColor="E2E8F0")
        fill_par = PatternFill("solid", fgColor=cor_linha_par)
        fill_impar = PatternFill("solid", fgColor=cor_linha_impar)

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        # =========================
        # TÃTULO DO RELATÃ“RIO
        # =========================
        total_cols = max(1, len(headers))
        ultima_coluna = get_column_letter(total_cols)
        col_widths: dict[int, int] = {}
        for i in range(1, total_cols + 1):
            hdr = headers[i - 1] if (i - 1) < len(headers) else ""
            col_widths[i] = min(max(len(str(hdr)) + 2, 12), 35)

        ws.merge_cells(f"A1:{ultima_coluna}1")
        ws["A1"] = "Pronto ClÃ­nica Arnaldo Quintela"
        ws["A1"].font = font_title
        ws["A1"].alignment = align_left

        ws.merge_cells(f"A2:{ultima_coluna}2")
        ws["A2"] = f"RelatÃ³rio exportado em {datetime.now().strftime('%d/%m/%Y Ã s %H:%M')}"
        ws["A2"].font = font_subtitle
        ws["A2"].alignment = align_left

        # =========================
        # CABEÃ‡ALHO DA TABELA
        # =========================
        header_row = 4
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.border = border_all
            cell.alignment = align_center

        # =========================
        # DADOS
        # =========================
        data_start_row = header_row + 1

        for row_idx, row_data in enumerate(rows, start=data_start_row):
            row_fill = fill_par if row_idx % 2 == 0 else fill_impar

            for col_idx, value in enumerate(row_data, start=1):
                safe_value = self._sanitize_spreadsheet_cell(value)
                cell = ws.cell(row=row_idx, column=col_idx, value=safe_value)
                cell.font = font_body
                cell.border = border_all
                cell.fill = row_fill

                try:
                    display_len = len("" if safe_value is None else str(safe_value))
                except Exception:
                    display_len = 0
                if display_len > 0:
                    col_widths[col_idx] = min(max(col_widths.get(col_idx, 12), display_len + 2), 35)

                header_name = str(headers[col_idx - 1]).strip().lower()

                if any(k in header_name for k in ["mat", "cpf", "status", "pagamento", "mÃªs", "mes", "data"]):
                    cell.alignment = align_center
                elif any(k in header_name for k in ["valor", "receita", "total"]):
                    cell.alignment = align_right
                else:
                    cell.alignment = align_left

                if any(k in header_name for k in ["valor", "receita", "total"]):
                    try:
                        if value not in (None, ""):
                            num = self._to_currency_number(value)
                            if num is not None:
                                cell.value = num
                                cell.number_format = 'R$ #,##0.00'
                    except Exception:
                        pass

        # =========================
        # LINHA DE TOTAL / RESUMO
        # =========================
        total_row = data_start_row + len(rows) + 1
        ws.cell(row=total_row, column=1, value="Total de registros")
        ws.cell(row=total_row, column=2, value=len(rows))

        ws.cell(row=total_row, column=1).font = font_total
        ws.cell(row=total_row, column=2).font = font_total
        ws.cell(row=total_row, column=1).fill = fill_total
        ws.cell(row=total_row, column=2).fill = fill_total
        ws.cell(row=total_row, column=1).border = border_all
        ws.cell(row=total_row, column=2).border = border_all
        ws.cell(row=total_row, column=1).alignment = align_left
        ws.cell(row=total_row, column=2).alignment = align_center

        # Se existir coluna de valor, soma automÃ¡tica
        col_valor = None
        for i, h in enumerate(headers, start=1):
            htxt = str(h).strip().lower()
            if "valor" in htxt:
                col_valor = i
                break

        if col_valor is not None and len(rows) > 0:
            soma_label_col = max(1, col_valor - 1)
            ws.cell(row=total_row + 1, column=soma_label_col, value="Soma dos valores")
            soma_cell = ws.cell(
                row=total_row + 1,
                column=col_valor,
                value=f"=SUM({get_column_letter(col_valor)}{data_start_row}:{get_column_letter(col_valor)}{data_start_row + len(rows) - 1})"
            )

            ws.cell(row=total_row + 1, column=soma_label_col).font = font_total
            ws.cell(row=total_row + 1, column=soma_label_col).fill = fill_total
            ws.cell(row=total_row + 1, column=soma_label_col).border = border_all
            ws.cell(row=total_row + 1, column=soma_label_col).alignment = align_right

            soma_cell.font = font_total
            soma_cell.fill = fill_total
            soma_cell.border = border_all
            soma_cell.alignment = align_right
            soma_cell.number_format = 'R$ #,##0.00'
            col_widths[soma_label_col] = min(max(col_widths.get(soma_label_col, 12), len("Soma dos valores") + 2), 35)

        # =========================
        # FILTRO E CONGELAR CABEÃ‡ALHO
        # =========================
        if len(rows) > 0:
            ws.auto_filter.ref = f"A{header_row}:{ultima_coluna}{header_row + len(rows)}"
        ws.freeze_panes = f"A{data_start_row}"

        # =========================
        # AJUSTE DE LARGURA
        # =========================
        for col_idx in range(1, total_cols + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_idx, 12)

        # =========================
        # ALTURA DAS LINHAS
        # =========================
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[header_row].height = 22

        wb.save(path)

    def _export(self, path: str, headers: list[str], rows: list[list]):
        if path.lower().endswith(".xlsx"):
            self._write_xlsx(path, headers, rows)
        else:
            if not path.lower().endswith(".csv"):
                path += ".csv"
            self._write_csv(path, headers, rows)

    def _record_activity(self, title: str, *, detail: str = "", level: str = "info", source: str = "system"):
        entry = ActivityEntry(
            when=datetime.now().strftime("%d/%m %H:%M"),
            title=str(title or "").strip() or "Atividade",
            detail=str(detail or "").strip(),
            level=str(level or "info").strip().lower(),
            source=str(source or "system").strip().lower(),
        ).to_dict()
        self._activity_history.insert(0, entry)
        self._activity_history = self._activity_history[:12]
        self._append_audit_log(entry)
        self._invalidate_dashboard_cache()

    def _append_audit_log(self, entry: dict):
        """Persist basic operational audit trail in JSONL (monthly files)."""
        try:
            now = datetime.now()
            audit_dir = db.get_app_data_dir() / "reports" / "auditoria"
            audit_dir.mkdir(parents=True, exist_ok=True)
            dst = audit_dir / f"auditoria_{now.strftime('%Y_%m')}.jsonl"
            payload = {
                "ts": now.strftime("%Y-%m-%d %H:%M:%S"),
                "role": str(self._nivel_usuario or "").strip() or "-",
                "title": str(entry.get("title", "") or ""),
                "detail": str(entry.get("detail", "") or ""),
                "level": str(entry.get("level", "") or "info"),
                "source": str(entry.get("source", "") or "system"),
            }
            with dst.open("a", encoding="utf-8") as f:
                f.write(json.dumps(payload, ensure_ascii=False) + "\n")
        except Exception as exc:
            logger.debug("Falha ao persistir auditoria operacional: %s", exc)

    def _write_renovacao_lote_report(
        self,
        *,
        stage: str,
        solicitados: list[int],
        atualizados: list[int],
        tipo_pdf: str = "",
        output_dir: str = "",
        pdf_generated: list[dict] | None = None,
        pdf_failed: list[dict] | None = None,
        cancelled: bool = False,
        extra: dict | None = None,
    ) -> dict:
        try:
            now = datetime.now()
            reports_dir = db.get_app_data_dir() / "reports" / "renovacoes"
            reports_dir.mkdir(parents=True, exist_ok=True)
            stamp = now.strftime("%Y%m%d_%H%M%S")
            stage_safe = str(stage or "geral").strip().lower().replace(" ", "_")
            base_name = f"renovacao_{stage_safe}_{stamp}"
            json_path = reports_dir / f"{base_name}.json"
            txt_path = reports_dir / f"{base_name}.txt"

            solicitados_set: set[int] = set()
            for raw in (solicitados or []):
                try:
                    value = int(raw)
                except Exception:
                    continue
                if value > 0:
                    solicitados_set.add(value)
            solicitados_clean = sorted(solicitados_set)

            atualizados_set: set[int] = set()
            for raw in (atualizados or []):
                try:
                    value = int(raw)
                except Exception:
                    continue
                if value > 0:
                    atualizados_set.add(value)
            atualizados_clean = sorted(atualizados_set)
            gen = list(pdf_generated or [])
            fail = list(pdf_failed or [])

            payload = {
                "gerado_em": now.strftime("%Y-%m-%d %H:%M:%S"),
                "usuario_nivel": str(self._nivel_usuario or "").strip() or "-",
                "stage": stage_safe,
                "cancelled": bool(cancelled),
                "solicitados": solicitados_clean,
                "atualizados": atualizados_clean,
                "qtd_solicitados": len(solicitados_clean),
                "qtd_atualizados": len(atualizados_clean),
                "tipo_pdf": str(tipo_pdf or "").strip().lower(),
                "output_dir": str(output_dir or "").strip(),
                "pdf_generated": gen,
                "pdf_failed": fail,
                "extra": dict(extra or {}),
            }
            json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

            lines = [
                "MEDCONTRACT - RELATORIO DE RENOVACAO EM LOTE",
                f"Gerado em: {payload['gerado_em']}",
                f"Usuario/Perfil: {payload['usuario_nivel']}",
                f"Etapa: {stage_safe}",
                f"Cancelado: {'SIM' if payload['cancelled'] else 'NAO'}",
                f"Solicitados: {payload['qtd_solicitados']}",
                f"Atualizados: {payload['qtd_atualizados']}",
            ]
            if payload["tipo_pdf"]:
                lines.append(f"Tipo PDF: {str(payload['tipo_pdf']).upper()}")
            if payload["output_dir"]:
                lines.append(f"Pasta destino: {payload['output_dir']}")
            if gen or fail:
                lines.append(f"PDFs gerados: {len(gen)}")
                lines.append(f"PDFs com falha: {len(fail)}")
            lines.append("")

            if atualizados_clean:
                lines.append("CLIENTES ATUALIZADOS (MAT):")
                lines.append(", ".join(str(v) for v in atualizados_clean))
                lines.append("")

            if gen:
                lines.append("PDFS GERADOS:")
                for item in gen:
                    mat = int(item.get("mat", 0) or 0)
                    nome = str(item.get("nome", "") or "").strip() or f"MAT {mat}"
                    pdf_name = Path(str(item.get("pdf_path", "") or "")).name
                    lines.append(f"- MAT {mat} | {nome} | {pdf_name}")
                lines.append("")

            if fail:
                lines.append("FALHAS:")
                for item in fail:
                    mat = int(item.get("mat", 0) or 0)
                    err = str(item.get("erro", "") or "").strip()
                    lines.append(f"- MAT {mat} | {err[:220]}")
                lines.append("")

            txt_path.write_text("\n".join(lines), encoding="utf-8")
            return {
                "ok": True,
                "json_path": str(json_path),
                "txt_path": str(txt_path),
                "name": json_path.name,
            }
        except Exception as exc:
            logger.warning("Falha ao gerar relatorio de renovacao em lote: %s", exc)
            return {"ok": False, "error": _sanitize_error_text(str(exc))}

    def _record_export_event(self, action: str, *, ok: bool, path: str = "", error: str = ""):
        when = datetime.now().strftime("%d/%m %H:%M")
        detail = ""
        if ok and path:
            detail = path.split("\\")[-1].split("/")[-1]
        elif not ok and error:
            detail = str(error).strip()

        entry = {
            "when": when,
            "action": action,
            "ok": bool(ok),
            "detail": detail,
        }
        self._export_history.insert(0, entry)
        self._export_history = self._export_history[:5]
        self._record_activity(
            f"Exportação: {action}",
            detail=detail or ("Concluída" if ok else "Falhou"),
            level=("success" if ok else "warn"),
            source="export",
        )

    def _auto_export_tick(self):
        if not _env_flag("MEDCONTRACT_AUTO_EXPORT_ENABLED", False):
            return
        try:
            run_hour = int((os.getenv("MEDCONTRACT_AUTO_EXPORT_HOUR") or "8").strip())
        except Exception:
            run_hour = 8
        run_hour = max(0, min(23, run_hour))

        now = datetime.now()
        if now.hour < run_hour:
            return

        day_key = now.strftime("%Y-%m-%d")
        if self._last_auto_export_key == day_key:
            return
        self._run_daily_recurring_export(now, day_key)

    def _run_daily_recurring_export(self, now: datetime, day_key: str):
        if not self._ensure_export_allowed(popup=False):
            return
        headers = ["MAT", "Nome", "CPF", "Status", "Pagamento", "Mês Ref.", "Data Pag.", "Valor Pago"]
        out = []
        page_size = 5000
        offset = 0
        while True:
            chunk = db.listar_clientes_export_ultimo_pagamento(
                limit=page_size,
                offset=offset,
                pagamento_status="atrasado",
            )
            if not chunk:
                break
            for r in chunk:
                mes_ref = str(r[5] if len(r) > 5 else "" or "").strip()
                data_pag = str(r[6] if len(r) > 6 else "" or "").strip()
                valor_pago = r[7] if len(r) > 7 else ""
                out.append(
                    [
                        r[0],
                        r[1],
                        r[2],
                        str(r[3] or "").replace("_", " ").upper(),
                        str(r[4] or "").replace("_", " ").upper(),
                        iso_to_mes_ref_br(mes_ref) if mes_ref else "",
                        self._iso_date_to_br(data_pag),
                        valor_pago,
                    ]
                )
            if len(chunk) < page_size:
                break
            offset += page_size

        if not out:
            self._last_auto_export_key = day_key
            return

        fmt = str(os.getenv("MEDCONTRACT_AUTO_EXPORT_FORMAT") or "xlsx").strip().lower()
        ext = "csv" if fmt == "csv" else "xlsx"
        base_dir = db.get_app_data_dir() / "reports" / "auto_exports"
        base_dir.mkdir(parents=True, exist_ok=True)
        dst = base_dir / f"inadimplentes_auto_{now.strftime('%Y%m%d_%H%M')}.{ext}"
        try:
            self._export(str(dst), headers, out)
            self._record_export_event("Autoexport inadimplentes", ok=True, path=str(dst))
            self._last_auto_export_key = day_key
            if hasattr(self.dashboard, "add_alert"):
                self.dashboard.add_alert("info", f"Autoexport concluído: {dst.name}")
        except Exception as exc:
            self._record_export_event("Autoexport inadimplentes", ok=False, error=str(exc))
            logger.warning("Falha no autoexport recorrente: %s", exc)

    def _auto_cobranca_tick(self):
        if not _env_flag("MEDCONTRACT_AUTO_COBRANCA_ENABLED", True):
            return
        if self._role() != ROLE_ADMIN:
            return
        if self._auto_cobranca_inflight:
            return
        try:
            run_hour = int((os.getenv("MEDCONTRACT_AUTO_COBRANCA_HOUR") or "9").strip())
        except Exception:
            run_hour = 9
        run_hour = max(0, min(23, run_hour))

        now = datetime.now()
        if now.hour < run_hour:
            return

        day_key = now.strftime("%Y-%m-%d")
        if self._last_auto_cobranca_key == day_key:
            return

        self._auto_cobranca_inflight = True
        worker = _Worker(self._run_daily_auto_cobranca, day_key)
        self._cobranca_workers.append(worker)

        def _cleanup():
            self._auto_cobranca_inflight = False
            try:
                self._cobranca_workers.remove(worker)
            except Exception:
                pass

        def _on_result(result):
            _cleanup()
            out = dict(result or {})
            if bool(out.get("executed", False)):
                self._last_auto_cobranca_key = day_key
            total = int(out.get("total_itens", 0) or 0)
            sent = int(out.get("emails_enviados", 0) or 0)
            failed = int(out.get("emails_falharam", 0) or 0)
            send_enabled = bool(out.get("send_email_enabled", False))
            report_name = str(out.get("report_name", "") or "").strip()

            if total <= 0:
                self._record_activity(
                    "Régua de cobrança: sem envios para hoje",
                    detail=report_name or "Sem pendências na régua D-3/D-1/D0/D+3.",
                    level="info",
                    source="automacao",
                )
                return

            if send_enabled:
                level = "success" if failed == 0 else "warn"
                detail = f"{sent} e-mail(s) enviado(s)"
                if failed > 0:
                    detail += f", {failed} falha(s)"
                if report_name:
                    detail += f" · {report_name}"
                self._record_activity(
                    "Régua de cobrança executada",
                    detail=detail,
                    level=level,
                    source="automacao",
                )
            else:
                detail = f"{total} lembrete(s) gerado(s) sem envio automático"
                if report_name:
                    detail += f" · {report_name}"
                self._record_activity(
                    "Régua de cobrança gerada",
                    detail=detail,
                    level="info",
                    source="automacao",
                )

        def _on_error(error_msg: str):
            _cleanup()
            msg = str(error_msg or "Falha na régua automática de cobrança.")
            self._record_activity(
                "Régua de cobrança falhou",
                detail=msg,
                level="warn",
                source="automacao",
            )
            logger.warning("Falha na régua de cobrança: %s", msg)

        worker.signals.result.connect(_on_result)
        worker.signals.error.connect(_on_error)
        self._thread_pool.start(worker)

    def _run_daily_auto_cobranca(self, day_key: str) -> dict:
        """Executa régua D-3/D-1/D0/D+3 para clientes e empresas."""
        try:
            ref_date = datetime.strptime(str(day_key or ""), "%Y-%m-%d").date()
        except Exception:
            ref_date = datetime.now().date()
            day_key = ref_date.strftime("%Y-%m-%d")

        mes_iso = ref_date.strftime("%Y-%m")
        mes_br = iso_to_mes_ref_br(mes_iso).upper()
        send_email_enabled = _env_flag("MEDCONTRACT_AUTO_COBRANCA_SEND_EMAIL", False)

        stage_map = {
            3: ("D-3", "Lembrete de vencimento (D-3)"),
            1: ("D-1", "Lembrete de vencimento (D-1)"),
            0: ("D0", "Vencimento hoje (D0)"),
            -3: ("D+3", "Cobrança pós-vencimento (D+3)"),
        }

        def _money_to_float(v) -> float:
            if v is None:
                return 0.0
            if isinstance(v, (int, float)):
                return float(v)
            txt = str(v).strip()
            if not txt:
                return 0.0
            txt = txt.replace("R$", "").replace("r$", "").replace(" ", "")
            if "," in txt and "." in txt:
                if txt.rfind(",") > txt.rfind("."):
                    txt = txt.replace(".", "").replace(",", ".")
                else:
                    txt = txt.replace(",", "")
            elif "," in txt:
                txt = txt.replace(".", "").replace(",", ".")
            try:
                return float(txt)
            except Exception:
                return 0.0

        def _wa_url(phone: str) -> str:
            digits = _only_digits(phone)
            if not digits:
                return ""
            if not digits.startswith("55"):
                digits = f"55{digits}"
            return f"https://wa.me/{digits}"

        def _build_msg(item: dict) -> tuple[str, str]:
            nome = str(item.get("nome", "Cliente") or "Cliente").strip() or "Cliente"
            valor = float(item.get("valor", 0.0) or 0.0)
            venc = str(item.get("vencimento_br", "") or "").strip()
            stage_title = str(item.get("stage_title", "Lembrete de cobrança") or "Lembrete de cobrança")
            status_legivel = "EM ATRASO" if str(item.get("status", "")).lower() == "em_atraso" else "PENDENTE"
            wa = str(item.get("whatsapp_url", "") or "").strip()
            subject = f"{stage_title} · {mes_br} · {nome}"
            lines = [
                f"Olá {nome},",
                "",
                f"Este é um aviso automático da sua mensalidade de {mes_br}.",
                f"Valor: {br_money(valor)}",
                f"Vencimento: {venc}",
                f"Situação atual: {status_legivel}",
                "",
                "Se o pagamento já foi realizado, por favor desconsidere este lembrete.",
            ]
            if wa:
                lines.append(f"WhatsApp para contato: {wa}")
            lines.extend(["", "MedContract · Automação de Cobrança"])
            return subject, "\n".join(lines)

        queue: list[dict] = []
        conn = None
        try:
            conn = db.connect()
            cur = conn.cursor()
            clientes_pagos_mes = db.cliente_ids_pagamento_mes_cursor(cur, mes_iso)
            empresas_pagas_mes = db.empresa_ids_pagamento_mes_cursor(cur, mes_iso)

            cur.execute(
                """
                SELECT
                    id, COALESCE(nome, ''), COALESCE(cpf, ''), COALESCE(telefone, ''), COALESCE(email, ''),
                    COALESCE(vencimento_dia, 10), COALESCE(valor_mensal, 0)
                FROM clientes
                WHERE COALESCE(status, 'ativo') <> 'inativo'
                """
            )
            for row in cur.fetchall() or []:
                cid = int(row[0] or 0)
                if cid in clientes_pagos_mes:
                    continue
                nome = str(row[1] or "").strip() or f"Cliente #{cid}"
                cpf = str(row[2] or "").strip()
                telefone = str(row[3] or "").strip()
                email = str(row[4] or "").strip()
                venc = max(1, min(31, int(row[5] or 10)))
                valor = max(0.0, float(row[6] or 0.0))
                dia = min(venc, int(monthrange(ref_date.year, ref_date.month)[1]))
                due = ref_date.replace(day=dia)
                delta = int((due - ref_date).days)
                if delta not in stage_map:
                    continue
                stage_key, stage_title = stage_map[delta]
                status = db.calcular_status_pagamento(
                    {"vencimento_dia": venc, "pagamento_mes_atual": False},
                    hoje=ref_date,
                )
                wa = _wa_url(telefone)
                queue.append(
                    {
                        "tipo": "cliente",
                        "id": cid,
                        "nome": nome,
                        "documento": cpf,
                        "telefone": telefone,
                        "email": email,
                        "valor": valor,
                        "vencimento": due.isoformat(),
                        "vencimento_br": due.strftime("%d/%m/%Y"),
                        "status": status,
                        "stage": stage_key,
                        "stage_title": stage_title,
                        "whatsapp_url": wa,
                    }
                )

            cur.execute(
                """
                SELECT
                    id, COALESCE(nome, ''), COALESCE(cnpj, ''), COALESCE(telefone, ''), COALESCE(email, ''),
                    COALESCE(dia_vencimento, 10), COALESCE(valor_mensal, '0')
                FROM empresas
                """
            )
            for row in cur.fetchall() or []:
                eid = int(row[0] or 0)
                if eid in empresas_pagas_mes:
                    continue
                nome = str(row[1] or "").strip() or f"Empresa #{eid}"
                cnpj = str(row[2] or "").strip()
                telefone = str(row[3] or "").strip()
                email = str(row[4] or "").strip()
                venc = max(1, min(31, int(row[5] or 10)))
                valor = max(0.0, _money_to_float(row[6]))
                dia = min(venc, int(monthrange(ref_date.year, ref_date.month)[1]))
                due = ref_date.replace(day=dia)
                delta = int((due - ref_date).days)
                if delta not in stage_map:
                    continue
                stage_key, stage_title = stage_map[delta]
                status = db.calcular_status_pagamento(
                    {"dia_vencimento": venc, "pagamento_mes_atual": False},
                    hoje=ref_date,
                )
                wa = _wa_url(telefone)
                queue.append(
                    {
                        "tipo": "empresa",
                        "id": eid,
                        "nome": nome,
                        "documento": cnpj,
                        "telefone": telefone,
                        "email": email,
                        "valor": valor,
                        "vencimento": due.isoformat(),
                        "vencimento_br": due.strftime("%d/%m/%Y"),
                        "status": status,
                        "stage": stage_key,
                        "stage_title": stage_title,
                        "whatsapp_url": wa,
                    }
                )
        finally:
            try:
                if conn:
                    conn.close()
            except Exception:
                pass

        queue.sort(key=lambda r: (str(r.get("stage", "")), str(r.get("nome", ""))))

        emails_enviados = 0
        emails_falharam = 0
        sem_email = 0

        if send_email_enabled:
            for item in queue:
                to_email = str(item.get("email", "") or "").strip()
                if not to_email or "@" not in to_email:
                    sem_email += 1
                    item["email_status"] = "sem_email"
                    continue
                subject, body = _build_msg(item)
                try:
                    email_service.send_email(to_email, subject, body)
                    emails_enviados += 1
                    item["email_status"] = "enviado"
                except Exception as exc:
                    emails_falharam += 1
                    item["email_status"] = "erro"
                    item["email_erro"] = _sanitize_error_text(str(exc))

        reports_dir = db.get_app_data_dir() / "reports" / "cobranca_automatica"
        reports_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        json_path = reports_dir / f"cobranca_{day_key}_{stamp}.json"
        txt_path = reports_dir / f"cobranca_{day_key}_{stamp}.txt"

        payload = {
            "gerado_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "dia_execucao": day_key,
            "mes_referencia": mes_iso,
            "send_email_enabled": bool(send_email_enabled),
            "total_itens": int(len(queue)),
            "emails_enviados": int(emails_enviados),
            "emails_falharam": int(emails_falharam),
            "sem_email": int(sem_email),
            "itens": queue,
        }
        json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

        lines = [
            "MEDCONTRACT - REGUA AUTOMATICA DE COBRANCA",
            f"Gerado em: {payload['gerado_em']}",
            f"Execucao: {day_key}",
            f"Mes referencia: {mes_iso}",
            f"Itens: {len(queue)}",
            f"E-mail automatico: {'SIM' if send_email_enabled else 'NAO'}",
            f"E-mails enviados: {emails_enviados}",
            f"E-mails falharam: {emails_falharam}",
            f"Sem e-mail: {sem_email}",
            "",
        ]
        for item in queue:
            lines.append(
                f"[{item.get('stage', '-')}] {item.get('tipo', '-').upper()} {item.get('nome', '-')}"
                f" | venc: {item.get('vencimento_br', '-')}"
                f" | valor: {br_money(float(item.get('valor', 0.0) or 0.0))}"
                f" | status: {str(item.get('status', '-') or '-').replace('_', ' ')}"
                f" | email: {item.get('email_status', 'n/a') if send_email_enabled else 'desativado'}"
            )
            wa = str(item.get("whatsapp_url", "") or "").strip()
            if wa:
                lines.append(f"  WhatsApp: {wa}")
        txt_path.write_text("\n".join(lines), encoding="utf-8")

        return {
            "executed": True,
            "day_key": day_key,
            "mes_ref": mes_iso,
            "total_itens": int(len(queue)),
            "emails_enviados": int(emails_enviados),
            "emails_falharam": int(emails_falharam),
            "sem_email": int(sem_email),
            "send_email_enabled": bool(send_email_enabled),
            "report_path": str(json_path),
            "report_name": json_path.name,
            "report_txt_path": str(txt_path),
        }

    def exportar_clientes(self):
        if not self._ensure_export_allowed():
            return
        try:
            headers = ["MAT", "Nome", "CPF", "Status", "Pagamento", "MÃªs Ref.", "Data Pag.", "Valor Pago"]
            out = []
            page_size = 5000
            offset = 0
            while True:
                chunk = db.listar_clientes_export_ultimo_pagamento(limit=page_size, offset=offset)
                if not chunk:
                    break

                for r in chunk:
                    mes_ref = str(r[5] if len(r) > 5 else "" or "").strip()
                    data_pag = str(r[6] if len(r) > 6 else "" or "").strip()
                    valor_pago = r[7] if len(r) > 7 else ""
                    out.append(
                        [
                            r[0],
                            r[1],
                            r[2],
                            str(r[3] or "").replace("_", " ").upper(),
                            str(r[4] or "").replace("_", " ").upper(),
                            iso_to_mes_ref_br(mes_ref) if mes_ref else "",
                            self._iso_date_to_br(data_pag),
                            valor_pago,
                        ]
                    )

                if len(chunk) < page_size:
                    break
                offset += page_size

            default = f"clientes_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            path = self._choose_export_path(default)
            if not path:
                return

            self._export(path, headers, out)
            self._record_export_event("Clientes", ok=True, path=path)
            QMessageBox.information(self, "ExportaÃ§Ã£o concluÃ­da", f"Arquivo salvo em:\n{path}")
            self.atualizar_dashboard_async()

        except Exception as e:
            safe_err = _sanitize_error_text(str(e))
            self._record_export_event("Clientes", ok=False, error=safe_err)
            QMessageBox.critical(self, "Falha na exportaÃ§Ã£o", safe_err)


    def exportar_inadimplentes(self):
        if not self._ensure_export_allowed():
            return
        try:
            headers = ["MAT", "Nome", "CPF", "Status", "Pagamento", "MÃªs Ref.", "Data Pag.", "Valor Pago"]

            out = []
            page_size = 5000
            offset = 0
            while True:
                chunk = db.listar_clientes_export_ultimo_pagamento(
                    limit=page_size,
                    offset=offset,
                    pagamento_status="atrasado",
                )
                if not chunk:
                    break

                for r in chunk:
                    mes_ref = str(r[5] if len(r) > 5 else "" or "").strip()
                    data_pag = str(r[6] if len(r) > 6 else "" or "").strip()
                    valor_pago = r[7] if len(r) > 7 else ""
                    out.append(
                        [
                            r[0],
                            r[1],
                            r[2],
                            str(r[3] or "").replace("_", " ").upper(),
                            str(r[4] or "").replace("_", " ").upper(),
                            iso_to_mes_ref_br(mes_ref) if mes_ref else "",
                            self._iso_date_to_br(data_pag),
                            valor_pago,
                        ]
                    )

                if len(chunk) < page_size:
                    break
                offset += page_size

            default = f"inadimplentes_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            path = self._choose_export_path(default)
            if not path:
                return

            self._export(path, headers, out)
            self._record_export_event("Inadimplentes", ok=True, path=path)
            QMessageBox.information(self, "ExportaÃ§Ã£o concluÃ­da", f"Arquivo salvo em:\n{path}")
            self.atualizar_dashboard_async()

        except Exception as e:
            safe_err = _sanitize_error_text(str(e))
            self._record_export_event("Inadimplentes", ok=False, error=safe_err)
            QMessageBox.critical(self, "Falha na exportaÃ§Ã£o", safe_err)


    def exportar_pagamentos_mes(self):
        if not self._ensure_export_allowed():
            return
        try:
            default_iso = datetime.now().strftime("%Y-%m")

            dlg = ExportPagamentosDialog(self, default_iso=default_iso)
            if dlg.exec() != QDialog.Accepted:
                return

            mes_iso = dlg.mes_iso()
            mes_br = iso_to_mes_ref_br(mes_iso)
            headers = ["MAT", "Nome", "CPF", "MÃªs Ref.", "Data Pag.", "Valor Pago"]

            out = []
            page_size = 5000
            offset = 0
            while True:
                chunk = db.listar_pagamentos_detalhados_mes(
                    mes_iso=mes_iso,
                    limit=page_size,
                    offset=offset,
                )
                if not chunk:
                    break

                for r in chunk:
                    out.append(
                        [
                            r.get("mat", ""),
                            r.get("nome", ""),
                            r.get("cpf", ""),
                            mes_br,
                            self._iso_date_to_br(r.get("data_pagamento", "")),
                            float(r.get("valor_pago", 0.0) or 0.0),
                        ]
                    )

                if len(chunk) < page_size:
                    break
                offset += page_size

            default = f"pagamentos_{mes_br.replace('/', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            path = self._choose_export_path(default)
            if not path:
                return

            self._export(path, headers, out)
            self._record_export_event(f"Pagamentos {mes_br}", ok=True, path=path)
            QMessageBox.information(self, "ExportaÃ§Ã£o concluÃ­da", f"Arquivo salvo em:\n{path}")
            self.atualizar_dashboard_async()

        except Exception as e:
            safe_err = _sanitize_error_text(str(e))
            self._record_export_event("Pagamentos", ok=False, error=safe_err)
            QMessageBox.critical(self, "Falha na exportaÃ§Ã£o", safe_err)

    def exportar_financeiro_filtrado(self, mes_iso: str, rows: list[dict], config: dict | None = None):
        if not self._ensure_export_allowed():
            return
        try:
            mes_ref = (mes_iso or "").strip()
            if len(mes_ref) != 7 or mes_ref[4] != "-":
                mes_ref = datetime.now().strftime("%Y-%m")
            cfg = dict(config or {})
            source = str(cfg.get("source", "receitas") or "receitas").strip().lower()

            if source == "contas_pagar":
                field_defs = [
                    ("vencimento", "Vencimento"),
                    ("descricao", "Descrição"),
                    ("categoria", "Categoria"),
                    ("fornecedor", "Fornecedor"),
                    ("forma_pagto", "Forma Pgto"),
                    ("status", "Status"),
                    ("valor", "Valor"),
                    ("data_pgto", "Data Pgto"),
                ]
                selected = [k for k, _ in field_defs if k in set(cfg.get("columns", []) or [])]
                if not selected:
                    selected = [k for k, _ in field_defs]
                header_map = {k: h for k, h in field_defs}
                query = dict(self.financeiro.current_contas_query() if hasattr(self.financeiro, "current_contas_query") else {})
                payload_rows: list[dict] = []
                page = 0
                page_size = 2000
                total = None
                while True:
                    chunk_payload = db.listar_contas_pagar_detalhado_payload(
                        mes_ref,
                        page=page,
                        limit=page_size,
                        search=str(query.get("search", "") or ""),
                        status=str(query.get("status", "") or ""),
                        categoria=str(query.get("categoria", "") or ""),
                        min_value=query.get("min_value"),
                        max_value=query.get("max_value"),
                        only_vencidas=bool(query.get("only_vencidas", False)),
                        vencem_hoje=bool(query.get("vencem_hoje", False)),
                        vencem_7d=bool(query.get("vencem_7d", False)),
                        sort_key=str(query.get("sort_key", "data_vencimento") or "data_vencimento"),
                        sort_dir=str(query.get("sort_dir", "asc") or "asc"),
                    ) or {}
                    chunk = list(chunk_payload.get("rows", []) or [])
                    payload_rows.extend(chunk)
                    if total is None:
                        total = int(chunk_payload.get("total", 0) or 0)
                    if not chunk:
                        break
                    if total is not None and len(payload_rows) >= total:
                        break
                    page += 1

                if not payload_rows:
                    payload_rows = list(rows or [])
                if not payload_rows:
                    QMessageBox.information(self, "Nada para exportar", "Nao ha registros no filtro atual.")
                    return

                mes_br = iso_to_mes_ref_br(mes_ref)
                headers = [header_map[k] for k in selected]
                out = []
                for r in payload_rows:
                    status = str(r.get("status", "") or "").replace("_", " ").upper()
                    data_pg = str(r.get("data_pagamento_real", "") or "")
                    if status != "PAGA":
                        data_pg = "—"
                    row_map = {
                        "vencimento": self._iso_date_to_br(r.get("data_vencimento", "")),
                        "descricao": str(r.get("descricao", "") or ""),
                        "categoria": str(r.get("categoria", "") or ""),
                        "fornecedor": str(r.get("fornecedor", "") or ""),
                        "forma_pagto": str(r.get("forma_pagamento", "") or ""),
                        "status": status,
                        "valor": float(r.get("valor_previsto", 0.0) or 0.0),
                        "data_pgto": self._iso_date_to_br(data_pg),
                    }
                    out.append([row_map[k] for k in selected])

                default = f"contas_pagar_{mes_ref.replace('-', '')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                path = self._choose_export_path(default)
                if not path:
                    return

                self._export(path, headers, out)
                self._record_export_event(f"Contas a pagar {mes_br}", ok=True, path=path)
                QMessageBox.information(self, "Exportacao concluida", f"Arquivo salvo em:\n{path}")
                self.atualizar_dashboard_async()
                self.atualizar_contas_pagar_async(mes_ref, force=True, query=query)
                return

            field_defs = [
                ("data", "Data Pag."),
                ("mes_ref", "Mes Ref."),
                ("mat", "MAT"),
                ("nome", "Nome"),
                ("cpf", "CPF"),
                ("status", "Status"),
                ("pag_status", "Pagamento"),
                ("valor", "Valor Pago"),
            ]
            selected = [k for k, _ in field_defs if k in set(cfg.get("columns", []) or [])]
            if not selected:
                selected = [k for k, _ in field_defs]
            header_map = {k: h for k, h in field_defs}
            query = self._normalize_finance_query(
                self.financeiro.current_query() if hasattr(self.financeiro, "current_query") else None
            )

            payload_rows: list[dict] = []
            page = 0
            page_size = 2000
            total = None
            while True:
                chunk_payload = db.listar_financeiro_detalhado_payload(
                    mes_ref,
                    page=page,
                    limit=page_size,
                    search_doc=str(query.get("search_doc", "") or ""),
                    search_name=str(query.get("search_name", "") or ""),
                    status_key=str(query.get("status_key", "") or ""),
                    min_value=query.get("min_value"),
                    max_value=query.get("max_value"),
                    only_atrasados=bool(query.get("only_atrasados", False)),
                    above_ticket=bool(query.get("above_ticket", False)),
                    ticket_ref=float(query.get("ticket_ref", 0.0) or 0.0),
                    only_today=bool(query.get("only_today", False)),
                    sort_key=str(query.get("sort_key", "data_pagamento") or "data_pagamento"),
                    sort_dir=str(query.get("sort_dir", "desc") or "desc"),
                ) or {}
                chunk = list(chunk_payload.get("rows", []) or [])
                payload_rows.extend(chunk)
                if total is None:
                    total = int(chunk_payload.get("total", 0) or 0)
                if not chunk:
                    break
                if total is not None and len(payload_rows) >= total:
                    break
                page += 1

            if not payload_rows:
                payload_rows = list(rows or [])
            if not payload_rows:
                QMessageBox.information(self, "Nada para exportar", "Nao ha registros no filtro atual.")
                return

            mes_br = iso_to_mes_ref_br(mes_ref)
            headers = [header_map[k] for k in selected]

            out = []
            for r in payload_rows:
                data_pag = str(r.get("data_pagamento", "") or "")
                if len(data_pag) == 10 and data_pag[4] == "-" and data_pag[7] == "-":
                    try:
                        y, m, d = data_pag.split("-")
                        data_br = f"{d}/{m}/{y}"
                    except Exception:
                        data_br = data_pag
                else:
                    data_br = data_pag

                row_mes = str(r.get("mes_referencia", "") or mes_ref)
                mes_ref_br = iso_to_mes_ref_br(row_mes) if row_mes else mes_br

                row_map = {
                    "data": data_br,
                    "mes_ref": mes_ref_br,
                    "mat": r.get("mat", ""),
                    "nome": str(r.get("nome", "") or ""),
                    "cpf": str(r.get("cpf", "") or ""),
                    "status": str(r.get("status", "") or "").replace("_", " ").upper(),
                    "pag_status": str(r.get("pagamento_status", "") or "").replace("_", " ").upper(),
                    "valor": float(r.get("valor_pago", 0.0) or 0.0),
                }
                out.append([row_map[k] for k in selected])

            default = f"financeiro_filtrado_{mes_ref.replace('-', '')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            path = self._choose_export_path(default)
            if not path:
                return

            self._export(path, headers, out)
            self._record_export_event(f"Financeiro {mes_br}", ok=True, path=path)
            QMessageBox.information(self, "Exportacao concluida", f"Arquivo salvo em:\n{path}")
            self.atualizar_dashboard_async()
        except Exception as e:
            safe_err = _sanitize_error_text(str(e))
            self._record_export_event("Financeiro filtrado", ok=False, error=safe_err)
            QMessageBox.critical(self, "Falha na exportacao", safe_err)

    def _handle_contas_pagar_action(self, action: dict):
        data = dict(action or {})
        kind = str(data.get("action", "") or "").strip().lower()

        try:
            if kind == "save":
                payload = dict(data.get("payload", {}) or {})
                conta_id = int(payload.get("id", 0) or 0)
                if conta_id > 0:
                    ok, msg, _ = db.atualizar_conta_pagar(conta_id, payload)
                else:
                    ok, msg, _ = db.salvar_conta_pagar(payload)
                if not ok:
                    if hasattr(self.financeiro, "show_contas_error"):
                        self.financeiro.show_contas_error(str(msg or "Não foi possível salvar a conta."))
                    QMessageBox.warning(self, "Conta a pagar", msg)
                    return
                self._invalidate_dashboard_cache()
                self._invalidate_finance_cache()
                self._invalidate_contas_cache()
                self.atualizar_dashboard_async(force=True)
                self.atualizar_contas_pagar_async(
                    self.financeiro.current_month() if hasattr(self.financeiro, "current_month") else None,
                    force=True,
                    query=(self.financeiro.current_contas_query() if hasattr(self.financeiro, "current_contas_query") else None),
                )
                if hasattr(self.financeiro, "show_contas_error"):
                    self.financeiro.show_contas_error("")
                QMessageBox.information(self, "Conta a pagar", msg or "Conta salva com sucesso.")
                return

            if kind == "pagar":
                conta_id = int(data.get("id", 0) or 0)
                if conta_id <= 0:
                    return
                ok, msg, _ = db.marcar_conta_paga(conta_id, data_pagamento_real=datetime.now().strftime("%Y-%m-%d"))
                if not ok:
                    if hasattr(self.financeiro, "show_contas_error"):
                        self.financeiro.show_contas_error(str(msg or "Não foi possível marcar a conta como paga."))
                    QMessageBox.warning(self, "Conta a pagar", msg)
                    return
                self._invalidate_dashboard_cache()
                self._invalidate_finance_cache()
                self._invalidate_contas_cache()
                self.atualizar_dashboard_async(force=True)
                self.atualizar_contas_pagar_async(
                    self.financeiro.current_month() if hasattr(self.financeiro, "current_month") else None,
                    force=True,
                    query=(self.financeiro.current_contas_query() if hasattr(self.financeiro, "current_contas_query") else None),
                )
                if hasattr(self.financeiro, "show_contas_error"):
                    self.financeiro.show_contas_error("")
                QMessageBox.information(self, "Conta a pagar", msg or "Conta marcada como paga.")
                return

            if kind == "delete":
                conta_id = int(data.get("id", 0) or 0)
                if conta_id <= 0:
                    return
                ok, msg = db.excluir_conta_pagar(conta_id)
                if not ok:
                    if hasattr(self.financeiro, "show_contas_error"):
                        self.financeiro.show_contas_error(str(msg or "Não foi possível excluir a conta."))
                    QMessageBox.warning(self, "Conta a pagar", msg)
                    return
                self._invalidate_dashboard_cache()
                self._invalidate_finance_cache()
                self._invalidate_contas_cache()
                self.atualizar_dashboard_async(force=True)
                self.atualizar_contas_pagar_async(
                    self.financeiro.current_month() if hasattr(self.financeiro, "current_month") else None,
                    force=True,
                    query=(self.financeiro.current_contas_query() if hasattr(self.financeiro, "current_contas_query") else None),
                )
                if hasattr(self.financeiro, "show_contas_error"):
                    self.financeiro.show_contas_error("")
                QMessageBox.information(self, "Conta a pagar", msg or "Conta excluída com sucesso.")
                return
        except Exception as e:
            if hasattr(self.financeiro, "show_contas_error"):
                self.financeiro.show_contas_error(f"Falha ao processar ação: {_sanitize_error_text(str(e))}")
            QMessageBox.critical(self, "Conta a pagar", f"Falha ao processar ação: {_sanitize_error_text(str(e))}")

    # ============================
    # Preview CPF (Pagamento)
    # ============================
    def pagamento_preview_por_cpf(self, cpf: str):
        try:
            info = db.buscar_cliente_preview_por_cpf(cpf)
            if not info:
                self.pagamento.set_cliente_preview(False, "CPF nao encontrado.")
                return

            status_l = (info.get("status") or "").lower()
            pag_l = (info.get("pagamento_status") or "").lower()

            status_txt = (info.get("status") or "-").upper()
            pag_txt = (info.get("pagamento_status") or "-").replace("_", " ").upper()

            plano = info.get("plano") or "-"
            deps = int(info.get("dependentes") or 0)
            deps_lista = info.get("dependentes_lista") or []
            vm = info.get("valor_mensal")
            ultimo = info.get("ultimo_pagamento")

            texto = f"Cliente: {info.get('nome','-')} - Plano: {plano} - Dep: {deps}\nStatus: {status_txt} - Pagamento: {pag_txt}"

            if ultimo:
                try:
                    texto += (
                        f"\nUltimo pagamento: {iso_to_mes_ref_br(ultimo.get('mes_referencia',''))} - "
                        f"{ultimo.get('data_pagamento','-')} - "
                        f"R$ {float(ultimo.get('valor_pago', 0.0)):.2f}"
                    )
                except Exception:
                    pass

            if deps_lista:
                nomes = ", ".join(d.get("nome", "-") for d in deps_lista[:3])
                if len(deps_lista) > 3:
                    nomes += f" +{len(deps_lista) - 3}"
                texto += f"\nDependentes: {nomes}"

            warn = (status_l == "inativo")
            if warn:
                texto += "  -  INATIVO"

            try:
                self.pagamento.set_cliente_preview(
                    True,
                    texto,
                    warn=warn,
                    cliente_id=info.get("id"),
                    nome=info.get("nome"),
                    status=status_l,
                    pagamento_status=pag_l,
                    plano=info.get("plano"),
                    dependentes=deps,
                    valor_mensal=vm,
                    ultimo_pagamento=ultimo,
                )
            except TypeError:
                self.pagamento.set_cliente_preview(
                    True,
                    texto,
                    warn=warn,
                    cliente_id=info.get("id"),
                    nome=info.get("nome"),
                    status=status_l,
                    pagamento_status=pag_l,
                    plano=info.get("plano"),
                    dependentes=deps,
                    valor_mensal=vm,
                )

        except Exception:
            self.pagamento.set_cliente_preview(False, "Erro ao consultar CPF.")

    def pagamento_preview_por_cnpj(self, cnpj: str):
        try:
            info = db.buscar_empresa_preview_por_cnpj(cnpj)
            if not info:
                self.pagamento.set_empresa_preview(False, "CNPJ nao encontrado.")
                return

            status_raw = (info.get("status_pagamento") or "").lower()
            status_txt = (info.get("status_pagamento") or "-").replace("_", " ").upper()
            forma_raw = (info.get("forma_pagamento") or "").lower()
            forma_txt = (info.get("forma_pagamento") or "-").replace("_", " ").upper()
            dia_venc = int(info.get("dia_vencimento") or 0)
            valor = info.get("valor_mensal")
            ultimo = info.get("ultimo_pagamento")

            texto = (
                f"Empresa: {info.get('nome', '-')} - Forma: {forma_txt}\n"
                f"Status de pagamento: {status_txt}"
            )
            if dia_venc > 0:
                texto += f" - Vencimento: dia {dia_venc}"

            if ultimo:
                try:
                    texto += (
                        f"\nUltimo pagamento: {iso_to_mes_ref_br(ultimo.get('mes_referencia', ''))} - "
                        f"{ultimo.get('data_pagamento', '-')} - "
                        f"R$ {float(ultimo.get('valor_pago', 0.0)):.2f}"
                    )
                except Exception:
                    pass

            self.pagamento.set_empresa_preview(
                True,
                texto,
                empresa_id=info.get("id"),
                nome=info.get("nome"),
                status_pagamento=status_raw,
                forma_pagamento=forma_raw,
                dia_vencimento=dia_venc,
                valor_mensal=valor,
                ultimo_pagamento=ultimo,
            )
        except Exception:
            self.pagamento.set_empresa_preview(False, "Erro ao consultar CNPJ.")

    def pagamento_existe_por_cliente_mes(self, cliente_id: int, mes_iso: str):
        try:
            return db.pagamento_existe(int(cliente_id), str(mes_iso))
        except Exception:
            return False, None

    def pagamento_existe_por_empresa_mes(self, empresa_id: int, mes_iso: str):
        try:
            return db.pagamento_empresa_existe(int(empresa_id), str(mes_iso))
        except Exception:
            return False, None

    @staticmethod
    def _contract_type_from_forma_pagamento(forma_pagamento: str) -> str:
        txt = str(forma_pagamento or "").strip().lower()
        plain = "".join(
            ch for ch in unicodedata.normalize("NFKD", txt)
            if not unicodedata.combining(ch)
        )
        plain = plain.replace("ç", "c").replace("ã", "a")
        if "pix" in plain:
            return "pix"
        if "recepcao" in plain or "recep" in plain:
            return "recepcao"
        return "boleto"

    def _load_contract_cliente_data(self, mat: int) -> tuple[dict, list[dict]]:
        row = db.buscar_cliente_por_id(int(mat))
        if not row:
            raise ValueError("Cliente não encontrado.")

        cliente = {
            "id": row[0],
            "matricula": row[0],
            "nome": row[1] if len(row) > 1 else "",
            "cpf": row[2] if len(row) > 2 else "",
            "telefone": row[3] if len(row) > 3 else "",
            "email": row[4] if len(row) > 4 else "",
            "data_inicio": row[5] if len(row) > 5 else "",
            "valor_mensal": row[6] if len(row) > 6 else 0.0,
            "status": row[7] if len(row) > 7 else "ativo",
            "pagamento_status": row[8] if len(row) > 8 else "em_dia",
            "observacoes": row[9] if len(row) > 9 else "",
            "data_nascimento": row[10] if len(row) > 10 else "",
            "cep": row[11] if len(row) > 11 else "",
            "endereco": row[12] if len(row) > 12 else "",
            "plano": row[13] if len(row) > 13 else "",
            "dependentes": row[14] if len(row) > 14 else 0,
            "vencimento_dia": row[15] if len(row) > 15 else 10,
            "forma_pagamento": row[16] if len(row) > 16 else "",
        }

        deps_rows = db.listar_dependentes(int(mat)) or []
        dependentes = []
        for r in deps_rows:
            dependentes.append({
                "nome": r[1] if len(r) > 1 else "",
                "cpf": r[2] if len(r) > 2 else "",
                "data_nascimento": r[4] if len(r) > 4 else "",
                "idade": r[3] if len(r) > 3 else 0,
            })

        return cliente, dependentes

    def _generate_contract_pdf_worker(
        self,
        cliente: dict,
        dependentes: list[dict],
        tipo: str,
        operation: str = "manual",
        output_dir: str | None = None,
    ) -> str:
        from services.contract_service import generate_contract_pdf

        out_dir_path = Path(output_dir).expanduser() if str(output_dir or "").strip() else None
        pdf_path = generate_contract_pdf(
            cliente=cliente,
            dependentes=dependentes,
            contract_type=tipo,
            operation=operation,
            output_dir=out_dir_path,
        )
        return str(pdf_path)

    def _generate_contracts_pdf_batch_worker(
        self,
        mats: list[int],
        tipo: str,
        operation: str = "renovacao_lote",
        output_dir: str | None = None,
    ) -> dict:
        from services.contract_service import generate_contract_pdf

        out_dir_path = Path(output_dir).expanduser() if str(output_dir or "").strip() else None
        ok_items: list[dict] = []
        failed_items: list[dict] = []

        for raw in (mats or []):
            try:
                mat = int(raw)
            except Exception:
                continue
            if mat <= 0:
                continue

            try:
                cliente, dependentes = self._load_contract_cliente_data(mat)
                pdf_path = generate_contract_pdf(
                    cliente=cliente,
                    dependentes=dependentes,
                    contract_type=str(tipo or "boleto").strip().lower(),
                    operation=operation,
                    output_dir=out_dir_path,
                )
                ok_items.append(
                    {
                        "mat": mat,
                        "nome": str(cliente.get("nome", "") or "").strip() or f"MAT {mat}",
                        "pdf_path": str(pdf_path),
                    }
                )
            except Exception as exc:
                failed_items.append(
                    {
                        "mat": mat,
                        "erro": _sanitize_error_text(str(exc)),
                    }
                )

        return {
            "requested": len([int(v) for v in (mats or []) if str(v).strip().isdigit()]),
            "generated": ok_items,
            "failed": failed_items,
            "tipo": str(tipo or "").strip().lower(),
            "output_dir": str(out_dir_path) if out_dir_path else "",
        }

    def _generate_contract_pdf_by_mat_worker(
        self,
        mat: int,
        tipo: str,
        operation: str = "renovacao_lote",
        output_dir: str | None = None,
    ) -> dict:
        mat_i = int(mat)
        cliente, dependentes = self._load_contract_cliente_data(mat_i)
        pdf_path = self._generate_contract_pdf_worker(
            cliente,
            dependentes,
            str(tipo or "boleto").strip().lower() or "boleto",
            operation,
            output_dir,
        )
        return {
            "mat": mat_i,
            "nome": str(cliente.get("nome", "") or "").strip() or f"MAT {mat_i}",
            "pdf_path": str(pdf_path),
        }

    def _run_contract_pdf_batch_with_progress(
        self,
        mats: list[int],
        tipo: str,
        output_dir: str | None = None,
        *,
        operation: str = "renovacao_lote",
    ):
        mats_clean: list[int] = []
        seen: set[int] = set()
        for raw in (mats or []):
            try:
                value = int(raw)
            except Exception:
                continue
            if value <= 0 or value in seen:
                continue
            seen.add(value)
            mats_clean.append(value)

        if not mats_clean:
            return

        total = len(mats_clean)
        progress = QProgressDialog(
            "Preparando geração em lote...",
            "Cancelar",
            0,
            total,
            self,
        )
        progress.setWindowTitle("Gerando contratos em lote")
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.setAutoClose(False)
        progress.setAutoReset(False)
        progress.setValue(0)
        progress.canceled.connect(
            lambda: progress.setLabelText("Cancelando... finalizando o item atual.")
        )
        progress.show()

        state = {
            "index": 0,
            "generated": [],
            "failed": [],
            "cancelled": False,
        }

        def _finish():
            qtd_ok = len(state["generated"])
            qtd_fail = len(state["failed"])
            processados = int(state["index"])
            restantes = max(0, total - processados)

            try:
                progress.setValue(total)
            except Exception:
                pass
            try:
                progress.close()
                progress.deleteLater()
            except Exception:
                pass

            if qtd_ok > 0:
                first_name = Path(str((state["generated"][0] or {}).get("pdf_path", "") or "")).name
                detail = f"{qtd_ok} PDF(s) em lote"
                if first_name:
                    detail = f"{detail} • ex.: {first_name}"
                self._record_activity(
                    "PDFs de contratos gerados em lote",
                    detail=detail,
                    level="success",
                    source="contrato",
                )
            if qtd_fail > 0:
                self._record_activity(
                    "Falha parcial na geracao de PDFs em lote",
                    detail=f"Falhas: {qtd_fail}",
                    level="warn",
                    source="contrato",
                )

            report_meta = self._write_renovacao_lote_report(
                stage="pdf",
                solicitados=list(mats_clean),
                atualizados=[int((item or {}).get("mat", 0) or 0) for item in state["generated"]],
                tipo_pdf=str(tipo),
                output_dir=str(output_dir or ""),
                pdf_generated=list(state["generated"]),
                pdf_failed=list(state["failed"]),
                cancelled=bool(state["cancelled"]),
                extra={
                    "processados": int(processados),
                    "restantes": int(restantes),
                    "operation": str(operation or "renovacao_lote"),
                },
            )
            if bool(report_meta.get("ok")):
                self._record_activity(
                    "Relatorio de renovacao/PDF salvo",
                    detail=str(report_meta.get("name") or "renovacao_pdf"),
                    level="success",
                    source="relatorio",
                )

            if hasattr(self.listar, "_show_message"):
                if qtd_fail == 0 and not state["cancelled"]:
                    self.listar._show_message(f"PDFs gerados com sucesso ({qtd_ok}/{total}).", ok=True, ms=2600)
                else:
                    self.listar._show_message(
                        f"PDFs gerados: {qtd_ok}. Falhas: {qtd_fail}. Restantes: {restantes}.",
                        ok=False,
                    )

            resumo = (
                f"Processados: {processados}/{total}\n"
                f"PDFs gerados: {qtd_ok}\n"
                f"Falhas: {qtd_fail}\n"
                f"Tipo: {str(tipo).upper()}\n"
                f"Pasta: {str(output_dir or '-').strip() or '-'}"
            )
            if state["cancelled"] and restantes > 0:
                resumo += f"\n\nGeração interrompida pelo usuário. Restantes: {restantes}."
            if qtd_fail > 0:
                err_preview = str((state["failed"][0] or {}).get("erro") or "").strip()
                if err_preview:
                    resumo += f"\n\nPrimeira falha: {err_preview[:160]}"
            if bool(report_meta.get("ok")):
                resumo += f"\n\nRelatorio: {str(report_meta.get('json_path') or '-')}"
            QMessageBox.information(self, "Geracao de contratos em lote", resumo)

        def _step():
            if progress.wasCanceled():
                state["cancelled"] = True
                _finish()
                return

            if int(state["index"]) >= total:
                _finish()
                return

            idx = int(state["index"])
            mat = int(mats_clean[idx])
            progress.setLabelText(f"Gerando contrato {idx + 1}/{total} (MAT {mat})...")

            def _on_item_result(payload: dict, current_mat: int = mat):
                data = dict(payload or {})
                if not data.get("mat"):
                    data["mat"] = int(current_mat)
                state["generated"].append(data)
                state["index"] = int(state["index"]) + 1
                progress.setValue(int(state["index"]))
                QTimer.singleShot(0, _step)

            def _on_item_error(error_msg: str, current_mat: int = mat):
                state["failed"].append(
                    {
                        "mat": int(current_mat),
                        "erro": _sanitize_error_text(str(error_msg)),
                    }
                )
                state["index"] = int(state["index"]) + 1
                progress.setValue(int(state["index"]))
                QTimer.singleShot(0, _step)

            self._start_tracked_worker(
                self._generate_contract_pdf_by_mat_worker,
                int(mat),
                str(tipo),
                operation,
                output_dir,
                bucket=self._contract_workers,
                on_result=_on_item_result,
                on_error=_on_item_error,
            )

        QTimer.singleShot(0, _step)

    def baixar_contrato_cliente_por_mat(
        self,
        mat: int,
        operation: str = "manual",
        *,
        ask_options: bool = True,
        contract_type: str | None = None,
        output_dir: str | None = None,
        show_popup_on_success: bool = True,
        show_popup_on_error: bool = True,
    ):
        try:
            mat_i = int(mat)
        except Exception:
            QMessageBox.warning(self, "Gerar contrato em PDF", "Matrícula inválida.")
            return

        try:
            cliente, dependentes = self._load_contract_cliente_data(mat_i)
        except Exception as e:
            msg = f"Não foi possível localizar o cliente MAT {mat_i}: {e}"
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message(msg, ok=False)
            if hasattr(self.cadastro, "_show_message"):
                self.cadastro._show_message(msg, ok=False)
            return

        default_tipo = self._contract_type_from_forma_pagamento(cliente.get("forma_pagamento", ""))
        tipo = str(contract_type or default_tipo or "boleto").strip().lower()
        final_output_dir = str(output_dir or "").strip()
        if ask_options:
            dlg = ContractTypeDialog(self, default_type=default_tipo)
            if dlg.exec() != QDialog.Accepted:
                return
            tipo = dlg.selected_type() or default_tipo
            final_output_dir = dlg.selected_output_dir()

        if hasattr(self.listar, "_show_message"):
            self.listar._show_message(
                f"Gerando contrato em PDF do MAT {mat_i}. Aguarde a confirmação do download.",
                ok=True,
                ms=2800,
            )
        if hasattr(self.cadastro, "_show_message"):
            self.cadastro._show_message(
                "Gerando contrato em PDF em segundo plano. Aguarde a confirmação do download.",
                ok=True,
            )

        def _on_result(pdf_path: str):
            self._record_activity(
                "Contrato em PDF gerado",
                detail=Path(pdf_path).name if str(pdf_path or "").strip() else f"MAT {mat_i}",
                level="success",
                source="contrato",
            )
            ok_msg = f"Contrato em PDF gerado com sucesso em:\n{pdf_path}"
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message("PDF do contrato gerado com sucesso.", ok=True, ms=2200)
            if hasattr(self.cadastro, "_show_message"):
                self.cadastro._show_message("PDF do contrato gerado com sucesso.", ok=True)
            if show_popup_on_success:
                QMessageBox.information(self, "Contrato em PDF gerado", ok_msg)

        def _on_error(error_msg: str):
            self._record_activity(
                "Falha ao gerar contrato em PDF",
                detail=str(error_msg or "")[:180],
                level="warn",
                source="contrato",
            )
            msg = f"Não foi possível gerar o contrato em PDF.\n\nDetalhes: {_sanitize_error_text(str(error_msg))}"
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message("Falha ao gerar contrato em PDF.", ok=False)
            if hasattr(self.cadastro, "_show_message"):
                self.cadastro._show_message("Falha ao gerar contrato em PDF.", ok=False)
            if show_popup_on_error:
                QMessageBox.critical(self, "Falha ao gerar contrato", msg)
        self._start_tracked_worker(
            self._generate_contract_pdf_worker,
            cliente,
            dependentes,
            tipo,
            operation,
            final_output_dir,
            bucket=self._contract_workers,
            on_result=_on_result,
            on_error=_on_error,
        )

    # ============================
    # DB: salvar cliente
    # ============================
    def _salvar_cliente_worker(self, dados: dict) -> dict:
        return clientes_service.salvar_cliente(dict(dados or {}))

    def salvar_cliente_no_banco(self, dados: dict):
        if self._cliente_save_inflight:
            if hasattr(self.cadastro, "_show_message"):
                self.cadastro._show_message("Aguarde o salvamento atual terminar.", ok=True)
            return
        is_create_mode = str((dados or {}).get("modo") or "create").strip().lower() == "create"

        def _ok(cliente_id: int | None = None):
            if hasattr(self.cadastro, "sucesso_salvo"):
                try:
                    self.cadastro.sucesso_salvo(cliente_id)
                except TypeError:
                    self.cadastro.sucesso_salvo()
            if hasattr(self.listar, "reload"):
                self.listar.reload()
            self._invalidate_dashboard_cache()
            self.atualizar_dashboard_async()
            if is_create_mode and cliente_id:
                self._run_pos_cadastro_automation(int(cliente_id))

        def _err(msg: str):
            if hasattr(self.cadastro, "erro_salvo"):
                self.cadastro.erro_salvo(msg)

        self._cliente_save_inflight = True
        try:
            if hasattr(self.cadastro, "_set_save_saving"):
                self.cadastro._set_save_saving()
        except Exception:
            pass

        def _finish():
            self._cliente_save_inflight = False

        def _on_result(result: dict):
            ok = bool((result or {}).get("ok"))
            if ok:
                _ok((result or {}).get("cliente_id"))
            else:
                _err(str((result or {}).get("msg") or "Nao foi possivel salvar."))

        def _on_error(error_msg: str):
            _err(f"Erro ao salvar: {error_msg}")
        self._start_tracked_worker(
            self._salvar_cliente_worker,
            dict(dados or {}),
            bucket=self._cliente_save_workers,
            on_result=_on_result,
            on_error=_on_error,
            on_finish=_finish,
        )

    def _run_pos_cadastro_automation(self, cliente_id: int):
        try:
            row = db.buscar_cliente_por_id(int(cliente_id))
        except Exception:
            row = None
        if not row:
            return

        nome = str(row[1] if len(row) > 1 else "").strip() or f"MAT {cliente_id}"
        email = str(row[4] if len(row) > 4 else "").strip()
        plano = str(row[13] if len(row) > 13 else "").strip()
        forma = str(row[16] if len(row) > 16 else "").strip()
        vencimento = str(row[15] if len(row) > 15 else "").strip()

        if not self._show_modern_question(
            title="Automação de pós cadastro",
            subtitle="Cliente salvo com sucesso.",
            details=(
                "Ao confirmar, o sistema irá executar automaticamente:\n"
                "• geração do contrato em PDF\n"
                "• envio de e-mail de confirmação"
            ),
            confirm_text="Iniciar automação",
            cancel_text="Agora não",
        ):
            return

        self.baixar_contrato_cliente_por_mat(
            int(cliente_id),
            operation="pos_cadastro",
            ask_options=False,
            show_popup_on_success=False,
            show_popup_on_error=False,
        )

        if not email or "@" not in email:
            if hasattr(self.cadastro, "_show_message"):
                self.cadastro._show_message(
                    "Automação iniciada: contrato em geração. E-mail não enviado (cliente sem e-mail válido).",
                    ok=False,
                    ms=2800,
                )
            return

        subject = "Confirmação de cadastro de contrato"
        body = (
            f"Olá {nome},\n\n"
            "Seu cadastro foi concluído com sucesso no MedContract.\n"
            f"Plano: {plano or '-'}\n"
            f"Forma de pagamento: {forma or '-'}\n"
            f"Dia de vencimento: {vencimento or '-'}\n\n"
            "Se precisar de suporte, responda este e-mail.\n\n"
            "Atenciosamente,\n"
            "Equipe MedContract"
        )
        self.enviar_email_cliente(
            {
                "to_email": email,
                "subject": subject,
                "body_text": body,
                "nome": nome,
                "_silent": True,
                "_feedback_target": "cadastro",
            }
        )
        if hasattr(self.cadastro, "_show_message"):
            self.cadastro._show_message(
                "Automação iniciada: contrato e e-mail estão sendo processados em segundo plano.",
                ok=True,
                ms=3400,
            )

    # ============================
    # DB: empresas
    # ============================
    def _set_empresa_save_busy(self, busy: bool, msg: str | None = None):
        try:
            if hasattr(self.cadastro_empresa, "set_save_busy"):
                self.cadastro_empresa.set_save_busy(bool(busy), msg)
                return
        except Exception:
            pass

        try:
            if hasattr(self.cadastro_empresa, "btn_salvar"):
                self.cadastro_empresa.btn_salvar.setEnabled(not bool(busy))
            if hasattr(self.cadastro_empresa, "btn_cancelar"):
                self.cadastro_empresa.btn_cancelar.setEnabled(not bool(busy))
            if msg and hasattr(self.cadastro_empresa, "_show_message"):
                self.cadastro_empresa._show_message(str(msg), ok=True)
        except Exception:
            pass

    def _salvar_empresa_worker(self, modo: str, dados: dict) -> dict:
        payload = dict(dados or {})
        payload["modo"] = str(modo or payload.get("modo") or "create").strip().lower()
        ok, msg = empresa_controller.salvar_empresa(payload)
        return {
            "ok": bool(ok),
            "msg": (msg or "Empresa cadastrada com sucesso.") if ok else (msg or "Nao foi possivel cadastrar a empresa."),
        }

    def salvar_empresa_no_banco(self, dados: dict):
        modo = str(dados.get("modo") or "create").strip().lower()
        if self._empresa_save_inflight:
            if hasattr(self.cadastro_empresa, "_show_message"):
                self.cadastro_empresa._show_message("Aguarde o salvamento atual terminar.", ok=True)
            return

        def _ok(msg: str):
            if hasattr(self.cadastro_empresa, "sucesso_salvo"):
                self.cadastro_empresa.sucesso_salvo(msg)
            if hasattr(self.listar_empresas, "_show_message"):
                self.listar_empresas._show_message(msg, ok=True)
            if hasattr(self.listar_empresas, "reload"):
                self.listar_empresas.reload()
            self._invalidate_dashboard_cache()
            self.atualizar_dashboard_async()

            confirm_text = str(msg or "Empresa cadastrada com sucesso.")
            should_go_list = self._show_modern_question(
                title="Cadastro da empresa concluído",
                subtitle=confirm_text,
                details=(
                    "Escolha a próxima ação:\n"
                    "• Ir para a listagem de empresas\n"
                    "• Permanecer no cadastro para continuar"
                ),
                confirm_text="Ir para listagem",
                cancel_text="Permanecer aqui",
            )
            if should_go_list:
                self.ir_para_listar_empresas()
                return

            if modo == "create" and hasattr(self.cadastro_empresa, "set_create_mode"):
                self.cadastro_empresa.set_create_mode()
            self.stack.setCurrentWidget(self.cadastro_empresa)

        def _err(msg: str):
            if hasattr(self.cadastro_empresa, "erro_salvo"):
                self.cadastro_empresa.erro_salvo(msg)

        self._empresa_save_inflight = True
        self._set_empresa_save_busy(True, "Salvando empresa em segundo plano...")

        worker = _Worker(self._salvar_empresa_worker, modo, dict(dados or {}))
        self._empresa_save_workers.append(worker)

        def _cleanup():
            self._empresa_save_inflight = False
            self._set_empresa_save_busy(False)
            try:
                self._empresa_save_workers.remove(worker)
            except Exception:
                pass

        def _on_result(result: dict):
            _cleanup()
            ok = bool((result or {}).get("ok"))
            msg = str((result or {}).get("msg") or "").strip()
            if ok:
                _ok(msg or "Operacao concluida com sucesso.")
            else:
                _err(msg or "Nao foi possivel salvar a empresa.")

        def _on_error(error_msg: str):
            _cleanup()
            _err(f"Erro ao salvar empresa: {error_msg}")

        worker.signals.result.connect(_on_result)
        worker.signals.error.connect(_on_error)
        self._thread_pool.start(worker)

    def editar_empresa_por_id(self, empresa_id: int):
        row = db.buscar_empresa_por_id(int(empresa_id))
        if not row:
            if hasattr(self.listar_empresas, "_show_message"):
                self.listar_empresas._show_message("Empresa nao encontrada.", ok=False)
            return

        empresa = {
            "id": row[0],
            "cnpj": row[1],
            "nome": row[2],
            "telefone": row[3],
            "email": row[4],
            "logradouro": row[5],
            "numero": row[6],
            "bairro": row[7],
            "cep": row[8],
            "cidade": row[9],
            "estado": row[10],
            "forma_pagamento": row[11],
            "status_pagamento": row[12],
            "dia_vencimento": row[13],
            "valor_mensal": row[14],
            "data_cadastro": row[15],
        }

        if hasattr(self.cadastro_empresa, "set_edit_mode"):
            self.cadastro_empresa.set_edit_mode(empresa)
        self.stack.setCurrentWidget(self.cadastro_empresa)

    def _excluir_empresa_worker(self, empresa_id: int) -> dict:
        ok = db.excluir_empresa(int(empresa_id))
        return {"ok": bool(ok)}

    def excluir_empresa_por_id(self, empresa_id: int):
        # RBAC (Recepção): bloqueio no controlador para impedir exclusão por chamada direta.
        if self._role() == ROLE_RECEPCAO:
            if hasattr(self.listar_empresas, "_show_message"):
                self.listar_empresas._show_message("Perfil de recepção não pode excluir empresas.", ok=False)
            else:
                self._notify_access_denied("Perfil de recepção não pode excluir empresas.", popup=True)
            return
        if self._empresa_delete_inflight:
            if hasattr(self.listar_empresas, "_show_message"):
                self.listar_empresas._show_message("Ja existe uma exclusao de empresa em andamento.", ok=False)
            return

        row = db.buscar_empresa_por_id(int(empresa_id))
        nome = row[2] if row and len(row) > 2 else f"ID {empresa_id}"
        resp = QMessageBox.question(
            self,
            "Confirmar exclusao",
            f"Deseja excluir a empresa '{nome}'?\n\nEssa acao nao pode ser desfeita.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if resp != QMessageBox.Yes:
            return

        self._empresa_delete_inflight = True
        if hasattr(self.listar_empresas, "_show_message"):
            self.listar_empresas._show_message("Excluindo empresa...", ok=True)

        worker = _Worker(self._excluir_empresa_worker, int(empresa_id))
        self._empresa_delete_workers.append(worker)

        def _cleanup():
            self._empresa_delete_inflight = False
            try:
                self._empresa_delete_workers.remove(worker)
            except Exception:
                pass

        def _on_result(result: dict):
            _cleanup()
            ok = bool((result or {}).get("ok"))
            if ok:
                if hasattr(self.listar_empresas, "_show_message"):
                    self.listar_empresas._show_message("Empresa excluida com sucesso.", ok=True)
                if hasattr(self.listar_empresas, "reload"):
                    self.listar_empresas.reload()
                self._invalidate_dashboard_cache()
                self.atualizar_dashboard_async()
            else:
                if hasattr(self.listar_empresas, "_show_message"):
                    self.listar_empresas._show_message("Nao foi possivel excluir a empresa.", ok=False)

        def _on_error(error_msg: str):
            _cleanup()
            if hasattr(self.listar_empresas, "_show_message"):
                self.listar_empresas._show_message(
                    f"Nao foi possivel excluir a empresa: {error_msg}",
                    ok=False,
                )

        worker.signals.result.connect(_on_result)
        worker.signals.error.connect(_on_error)
        self._thread_pool.start(worker)

    # ============================
    # DB: editar/excluir
    # ============================
    def editar_cliente_por_mat(self, mat: int):
        if not self._can_edit_cliente():
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message("Perfil de recepção não pode editar clientes.", ok=False)
            else:
                self._notify_access_denied("Perfil de recepção não pode editar clientes.", popup=True)
            return

        row = db.buscar_cliente_por_id(int(mat))
        if not row:
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message("Cliente nÃ£o encontrado.", ok=False)
            return

        cliente = {
            "id": row[0],
            "matricula": row[0],
            "nome": row[1],
            "cpf": row[2],
            "telefone": row[3],
            "email": row[4],
            "data_inicio": row[5] if len(row) > 5 else "",
            "valor_mensal": row[6] if len(row) > 6 else 0.0,
            "status": row[7] if len(row) > 7 else "ativo",
            "pagamento_status": row[8] if len(row) > 8 else "em_dia",
            "observacoes": row[9] if len(row) > 9 else "",
            "data_nascimento": row[10] if len(row) > 10 else "",
            "cep": row[11] if len(row) > 11 else "",
            "endereco": row[12] if len(row) > 12 else "",
            "plano": row[13] if len(row) > 13 else "",
            "dependentes": row[14] if len(row) > 14 else 0,
            "vencimento_dia": row[15] if len(row) > 15 else 10,
            "forma_pagamento": row[16] if len(row) > 16 else "",
        }

        self.cadastro.set_edit_mode(cliente)

        try:
            deps_rows = db.listar_dependentes(int(mat))
            deps = [
                {
                    "nome": r[1],
                    "cpf": r[2],
                    "idade": r[3],
                    "data_nascimento": r[4] if len(r) > 4 else "",
                }
                for r in deps_rows
            ]
            self.cadastro.set_dependentes_lista(deps)
        except Exception:
            self.cadastro.set_dependentes_lista([])

        self.stack.setCurrentWidget(self.cadastro)

    def _excluir_cliente_worker(self, mat: int) -> dict:
        return clientes_service.excluir_cliente(int(mat))

    def excluir_cliente_por_mat(self, mat: int):
        # RBAC (Recepção): bloqueio no controlador para impedir exclusão por chamada direta.
        if self._role() == ROLE_RECEPCAO:
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message("Perfil de recepção não pode excluir clientes.", ok=False)
            else:
                self._notify_access_denied("Perfil de recepção não pode excluir clientes.", popup=True)
            return
        if self._cliente_delete_inflight:
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message("Ja existe uma exclusao de cliente em andamento.", ok=False)
            return

        resp = QMessageBox.question(
            self,
            "Confirmar exclusÃ£o",
            f"Tem certeza que deseja excluir o cliente MAT {mat}?\n\n"
            f"Isso removerÃ¡ tambÃ©m os pagamentos e dependentes deste cliente.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if resp != QMessageBox.Yes:
            return

        self._cliente_delete_inflight = True
        if hasattr(self.listar, "_show_message"):
            self.listar._show_message("Excluindo cliente...", ok=True)

        worker = _Worker(self._excluir_cliente_worker, int(mat))
        self._cliente_delete_workers.append(worker)

        def _cleanup():
            self._cliente_delete_inflight = False
            try:
                self._cliente_delete_workers.remove(worker)
            except Exception:
                pass

        def _on_result(result: dict):
            _cleanup()
            ok = bool((result or {}).get("ok"))
            if ok:
                if hasattr(self.listar, "_show_message"):
                    self.listar._show_message("Cliente excluido com sucesso.", ok=True)
                if hasattr(self.listar, "reload"):
                    self.listar.reload()
                self._invalidate_dashboard_cache()
                self.atualizar_dashboard_async()
            else:
                if hasattr(self.listar, "_show_message"):
                    self.listar._show_message("Nao foi possivel excluir.", ok=False)

        def _on_error(error_msg: str):
            _cleanup()
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message(
                    f"Nao foi possivel excluir: {error_msg}",
                    ok=False,
                )

        worker.signals.result.connect(_on_result)
        worker.signals.error.connect(_on_error)
        self._thread_pool.start(worker)

    def _cancelar_plano_worker(self, mat: int) -> dict:
        return clientes_service.cancelar_plano_cliente(int(mat))

    def cancelar_plano_cliente_por_mat(self, mat: int):
        if not self._can_edit_cliente():
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message("Perfil de recepcao nao pode editar clientes.", ok=False)
            else:
                self._notify_access_denied("Perfil de recepcao nao pode editar clientes.", popup=True)
            return
        if self._cancelar_plano_inflight:
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message("Ja existe um cancelamento em andamento.", ok=False)
            return

        row = db.buscar_cliente_por_id(int(mat))
        if not row:
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message("Cliente nao encontrado.", ok=False)
            return

        status_atual = str(row[7] if len(row) > 7 else "").strip().lower()
        if status_atual == "inativo":
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message("Cliente ja esta com plano cancelado.", ok=False)
            return

        self._cancelar_plano_inflight = True
        if hasattr(self.listar, "_show_message"):
            self.listar._show_message("Cancelando plano...", ok=True)

        worker = _Worker(self._cancelar_plano_worker, int(mat))
        self._cancelar_plano_workers.append(worker)

        def _cleanup():
            self._cancelar_plano_inflight = False
            try:
                self._cancelar_plano_workers.remove(worker)
            except Exception:
                pass

        def _on_result(result: dict):
            _cleanup()
            ok = bool((result or {}).get("ok"))
            msg = str((result or {}).get("msg") or "")
            if ok:
                if hasattr(self.listar, "_show_message"):
                    self.listar._show_message(msg or "Plano cancelado com sucesso.", ok=True)
                if hasattr(self.listar, "reload"):
                    self.listar.reload()
                self._invalidate_dashboard_cache()
                self.atualizar_dashboard_async()
            else:
                if hasattr(self.listar, "_show_message"):
                    self.listar._show_message(msg or "Nao foi possivel cancelar o plano.", ok=False)

        def _on_error(error_msg: str):
            _cleanup()
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message(
                    f"Nao foi possivel cancelar o plano: {error_msg}",
                    ok=False,
                )

        worker.signals.result.connect(_on_result)
        worker.signals.error.connect(_on_error)
        self._thread_pool.start(worker)

    def _renovar_contrato_worker(self, mat: int) -> dict:
        return clientes_service.renovar_contrato_cliente(int(mat))

    def renovar_contrato_cliente_por_mat(self, mat: int):
        if not self._can_edit_cliente():
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message("Perfil de recepcao nao pode editar clientes.", ok=False)
            else:
                self._notify_access_denied("Perfil de recepcao nao pode editar clientes.", popup=True)
            return
        if self._renovar_contrato_inflight:
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message("Ja existe uma renovacao em andamento.", ok=False)
            return

        row = db.buscar_cliente_por_id(int(mat))
        if not row:
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message("Cliente nao encontrado.", ok=False)
            return
        nome = str(row[1] if len(row) > 1 else "").strip() or f"MAT {int(mat)}"

        self._renovar_contrato_inflight = True
        if hasattr(self.listar, "_show_message"):
            self.listar._show_message("Renovando contrato...", ok=True)

        worker = _Worker(self._renovar_contrato_worker, int(mat))
        self._renovar_contrato_workers.append(worker)

        def _cleanup():
            self._renovar_contrato_inflight = False
            try:
                self._renovar_contrato_workers.remove(worker)
            except Exception:
                pass

        def _on_result(result: dict):
            _cleanup()
            ok = bool((result or {}).get("ok"))
            msg = str((result or {}).get("msg") or "")
            info = dict((result or {}).get("info") or {})

            if not ok:
                fail_msg = msg or "Nao foi possivel renovar o contrato."
                if hasattr(self.listar, "_show_message"):
                    self.listar._show_message(fail_msg, ok=False)
                self._record_activity(
                    "Falha na renovacao de contrato",
                    detail=f"{nome} (MAT {int(mat)})",
                    level="warn",
                    source="clientes",
                )
                return

            if hasattr(self.listar, "_show_message"):
                self.listar._show_message(msg or "Contrato renovado com sucesso.", ok=True)
            if hasattr(self.listar, "reload"):
                self.listar.reload()

            data_inicio_nova = str(info.get("data_inicio_nova") or "").strip()
            detail = f"{nome} (MAT {int(mat)})"
            if data_inicio_nova:
                detail = f"{detail} - inicio {data_inicio_nova}"
            self._record_activity(
                "Contrato renovado",
                detail=detail,
                level="success",
                source="clientes",
            )

            self._invalidate_dashboard_cache()
            self.atualizar_dashboard_async()

            ask_pdf = QMessageBox.question(
                self,
                "Renovacao concluida",
                f"Contrato renovado para {nome}.\n\nDeseja gerar o novo PDF agora?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes,
            )
            if ask_pdf == QMessageBox.Yes:
                self.baixar_contrato_cliente_por_mat(int(mat), operation="renovacao")

        def _on_error(error_msg: str):
            _cleanup()
            message = f"Nao foi possivel renovar o contrato: {error_msg}"
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message(message, ok=False)
            self._record_activity(
                "Falha na renovacao de contrato",
                detail=f"{nome} (MAT {int(mat)})",
                level="warn",
                source="clientes",
            )

        worker.signals.result.connect(_on_result)
        worker.signals.error.connect(_on_error)
        self._thread_pool.start(worker)

    def _renovar_contratos_lote_worker(self, mats: list[int]) -> dict:
        return clientes_service.renovar_contratos_clientes(list(mats or []))

    def renovar_contratos_marcados(self, mats: list[int]):
        if not self._can_edit_cliente():
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message("Perfil de recepcao nao pode editar clientes.", ok=False)
            else:
                self._notify_access_denied("Perfil de recepcao nao pode editar clientes.", popup=True)
            return

        ids_set: set[int] = set()
        for raw in (mats or []):
            try:
                value = int(raw)
            except Exception:
                continue
            if value > 0:
                ids_set.add(value)
        ids = sorted(ids_set)
        if not ids:
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message("Marque ao menos um cliente para renovar.", ok=False)
            return
        if self._renovar_lote_inflight:
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message("Ja existe uma renovacao em lote em andamento.", ok=False)
            return

        self._renovar_lote_inflight = True
        if hasattr(self.listar, "_show_message"):
            self.listar._show_message(
                f"Renovando {len(ids)} contrato(s) em lote...",
                ok=True,
            )

        worker = _Worker(self._renovar_contratos_lote_worker, ids)
        self._renovar_lote_workers.append(worker)

        def _cleanup():
            self._renovar_lote_inflight = False
            try:
                self._renovar_lote_workers.remove(worker)
            except Exception:
                pass

        def _on_result(result: dict):
            _cleanup()
            ok = bool((result or {}).get("ok"))
            msg = str((result or {}).get("msg") or "")
            info = dict((result or {}).get("info") or {})
            qtd = int(info.get("clientes_atualizados", 0) or 0)
            atualizados_ids = []
            for raw in (info.get("cliente_ids") or ids):
                try:
                    value = int(raw)
                except Exception:
                    continue
                if value > 0:
                    atualizados_ids.append(value)
            atualizados_ids = sorted(set(atualizados_ids))

            if not ok:
                fail_msg = msg or "Nao foi possivel renovar os contratos marcados."
                if hasattr(self.listar, "_show_message"):
                    self.listar._show_message(fail_msg, ok=False)
                self._record_activity(
                    "Falha na renovacao em lote",
                    detail=f"Solicitados: {len(ids)}",
                    level="warn",
                    source="clientes",
                )
                return

            if hasattr(self.listar, "_show_message"):
                self.listar._show_message(msg or f"Renovacao em lote concluida ({qtd}).", ok=True)
            if hasattr(self.listar, "_clear_marcados"):
                try:
                    self.listar._clear_marcados()
                except Exception:
                    pass
            if hasattr(self.listar, "reload"):
                self.listar.reload()

            self._record_activity(
                "Renovacao em lote concluida",
                detail=f"Atualizados: {qtd} cliente(s)",
                level="success",
                source="clientes",
            )
            report_meta = self._write_renovacao_lote_report(
                stage="status",
                solicitados=list(ids),
                atualizados=list(atualizados_ids),
                extra={"qtd_atualizados": int(qtd)},
            )
            if bool(report_meta.get("ok")):
                self._record_activity(
                    "Relatorio de renovacao salvo",
                    detail=str(report_meta.get("name") or "renovacao_status"),
                    level="success",
                    source="relatorio",
                )
            self._invalidate_dashboard_cache()
            self.atualizar_dashboard_async()

            if qtd <= 0 or not atualizados_ids:
                return

            ask_pdf = QMessageBox.question(
                self,
                "Renovacao em lote concluida",
                f"{qtd} contrato(s) renovado(s).\n\nDeseja gerar os PDFs agora?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes,
            )
            if ask_pdf != QMessageBox.Yes:
                return

            default_tipo = "boleto"
            try:
                first_row = db.buscar_cliente_por_id(int(atualizados_ids[0]))
            except Exception:
                first_row = None
            if first_row:
                try:
                    default_tipo = self._contract_type_from_forma_pagamento(first_row[16] if len(first_row) > 16 else "")
                except Exception:
                    default_tipo = "boleto"

            dlg = ContractTypeDialog(self, default_type=default_tipo)
            if dlg.exec() != QDialog.Accepted:
                if hasattr(self.listar, "_show_message"):
                    self.listar._show_message("Geracao de PDFs cancelada.", ok=False)
                return

            tipo = dlg.selected_type() or default_tipo
            output_dir = dlg.selected_output_dir()
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message(
                    f"Gerando PDFs de {len(atualizados_ids)} contrato(s) renovado(s)...",
                    ok=True,
                    ms=2600,
                )
            self._run_contract_pdf_batch_with_progress(
                list(atualizados_ids),
                str(tipo),
                output_dir,
                operation="renovacao_lote",
            )

        def _on_error(error_msg: str):
            _cleanup()
            message = f"Nao foi possivel renovar em lote: {error_msg}"
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message(message, ok=False)
            self._record_activity(
                "Falha na renovacao em lote",
                detail=f"Solicitados: {len(ids)}",
                level="warn",
                source="clientes",
            )

        worker.signals.result.connect(_on_result)
        worker.signals.error.connect(_on_error)
        self._thread_pool.start(worker)

    def _aplicar_reajuste_worker(self, payload: dict) -> dict:
        return clientes_service.aplicar_reajuste(dict(payload or {}))

    @staticmethod
    def _br_money(v) -> str:
        try:
            s = f"{float(v or 0.0):,.2f}"
            s = s.replace(",", "X").replace(".", ",").replace("X", ".")
            return f"R$ {s}"
        except Exception:
            return "R$ 0,00"

    def _build_reajuste_detail_text(self, modo: str, msg: str, info: dict) -> str:
        qtd = int(info.get("clientes_atualizados", 0) or 0)
        soma_atual = float(info.get("soma_atual", 0.0) or 0.0)
        soma_nova = float(info.get("soma_reajustada", 0.0) or 0.0)
        detail_lines = [str(msg or "Reajuste aplicado com sucesso.")]

        if modo == "individual":
            nome = str(info.get("cliente_nome", "-") or "-")
            mat = int(info.get("cliente_id", 0) or 0)
            pct_est = float(info.get("percentual_estimado", 0.0) or 0.0)
            detail_lines.extend([
                f"Cliente: {nome} (MAT {mat})",
                f"Valor mensal: {self._br_money(soma_atual)} -> {self._br_money(soma_nova)}",
                f"Variação: {self._br_money(float(info.get('diferenca_total', 0.0) or 0.0))} ({pct_est:.2f}%)",
            ])
        elif modo == "selecionados":
            solicitados = int(info.get("clientes_solicitados", 0) or 0)
            detail_lines.extend([
                f"Clientes selecionados: {solicitados}",
                f"Clientes afetados: {qtd}",
                f"Soma dos valores: {self._br_money(soma_atual)} -> {self._br_money(soma_nova)}",
            ])
        else:
            plano_label = str(info.get("plano_label", "Todos os planos") or "Todos os planos")
            detail_lines.extend([
                f"Plano alvo: {plano_label}",
                f"Clientes afetados: {qtd}",
                f"Soma dos valores: {self._br_money(soma_atual)} -> {self._br_money(soma_nova)}",
                "Valores-base dos planos: inalterados.",
            ])
        return "\n".join(detail_lines)

    def aplicar_reajuste_planos(self, payload: dict):
        if not self._can_edit_cliente():
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message("Perfil de recepcao nao pode reajustar planos.", ok=False)
            else:
                self._notify_access_denied("Perfil de recepcao nao pode reajustar planos.", popup=True)
            return
        if self._reajuste_inflight:
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message("Ja existe um reajuste em andamento.", ok=False)
            return

        self._reajuste_inflight = True
        if hasattr(self.listar, "_show_message"):
            self.listar._show_message("Aplicando reajuste...", ok=True)

        worker = _Worker(self._aplicar_reajuste_worker, dict(payload or {}))
        self._reajuste_workers.append(worker)

        def _cleanup():
            self._reajuste_inflight = False
            try:
                self._reajuste_workers.remove(worker)
            except Exception:
                pass

        def _on_result(result: dict):
            _cleanup()
            ok = bool((result or {}).get("ok"))
            msg = str((result or {}).get("msg") or "")
            info = dict((result or {}).get("info", {}) or {})
            modo = str((result or {}).get("modo") or "filtros").strip().lower()

            if not ok:
                if hasattr(self.listar, "_show_message"):
                    self.listar._show_message(msg or "Nao foi possivel aplicar reajuste.", ok=False)
                QMessageBox.critical(self, "Falha no reajuste", msg or "Nao foi possivel aplicar reajuste.")
                return

            detail_text = self._build_reajuste_detail_text(modo, msg, info)
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message("Reajuste aplicado com sucesso.", ok=True, ms=2600)
            if hasattr(self.listar, "reload"):
                self.listar.reload()

            self._invalidate_dashboard_cache()
            self.atualizar_dashboard_async()
            QMessageBox.information(self, "Reajuste aplicado", detail_text)

        def _on_error(error_msg: str):
            _cleanup()
            message = f"Falha no reajuste: {error_msg}"
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message(message, ok=False)
            QMessageBox.critical(self, "Falha no reajuste", message)

        worker.signals.result.connect(_on_result)
        worker.signals.error.connect(_on_error)
        self._thread_pool.start(worker)

    def _send_email_with_retry_worker(self, to_email: str, subject: str, body_text: str) -> dict:
        last_error: Exception | None = None
        for _attempt in (1, 2):
            try:
                return email_service.send_email(to_email, subject, body_text)
            except Exception as exc:
                last_error = exc
        raise RuntimeError(f"Falha no envio após nova tentativa: {last_error}")

    def enviar_email_cliente(self, payload: dict):
        # RBAC (Recepção): bloqueio no controlador para impedir envio por chamada direta.
        if self._role() == ROLE_RECEPCAO:
            if hasattr(self.listar, "_show_message"):
                self.listar._show_message("Perfil de recepção não pode enviar e-mails por esta tela.", ok=False)
            else:
                self._notify_access_denied("Perfil de recepção não pode enviar e-mails por esta tela.", popup=True)
            return

        data = dict(payload or {})
        to_email = str(data.get("to_email", "") or "").strip()
        subject = str(data.get("subject", "") or "").strip()
        body_text = str(data.get("body_text", "") or "").strip()
        nome = str(data.get("nome", "") or "").strip()
        silent = bool(data.get("_silent"))
        feedback_target = str(data.get("_feedback_target", "listar") or "listar").strip().lower()

        def _notify(text: str, ok: bool):
            msg = str(text or "").strip()
            if silent and ok:
                return
            if feedback_target == "cadastro" and hasattr(self.cadastro, "_show_message"):
                self.cadastro._show_message(msg, ok=ok)
            elif hasattr(self.listar, "_show_message"):
                self.listar._show_message(msg, ok=ok)
            elif hasattr(self.cadastro, "_show_message"):
                self.cadastro._show_message(msg, ok=ok)

        try:
            email_service.validate_runtime_smtp_config()
        except Exception as e:
            msg = f"Configuracao SMTP invalida: {e}"
            _notify(msg, ok=False)
            if not silent:
                QMessageBox.warning(self, "Falha na configuracao", msg)
            return

        if not to_email or "@" not in to_email:
            msg = "E-mail do cliente invalido."
            _notify(msg, ok=False)
            if not silent:
                QMessageBox.warning(self, "E-mail invalido", msg)
            return
        if not subject:
            msg = "Assunto do e-mail nao informado."
            _notify(msg, ok=False)
            if not silent:
                QMessageBox.warning(self, "Dados incompletos", msg)
            return
        if not body_text:
            msg = "Mensagem do e-mail nao informada."
            _notify(msg, ok=False)
            if not silent:
                QMessageBox.warning(self, "Dados incompletos", msg)
            return

        target = f"{nome} <{to_email}>" if nome else to_email
        _notify(f"Enviando e-mail para {target}...", ok=True)
        logger.info("Iniciando envio de e-mail para %s", target)

        worker = _Worker(self._send_email_with_retry_worker, to_email, subject, body_text)
        self._email_workers.append(worker)

        def _cleanup_worker():
            try:
                self._email_workers.remove(worker)
            except Exception:
                pass

        def _on_result(_result):
            _cleanup_worker()
            _notify(f"E-mail enviado para {target}.", ok=True)
            logger.info("E-mail enviado com sucesso para %s", target)

        def _on_error(error_msg: str):
            _cleanup_worker()
            message = str(error_msg or "Falha ao enviar e-mail.")
            _notify(message, ok=False)
            if not silent:
                QMessageBox.critical(self, "Falha no envio de e-mail", message)
            logger.error("Falha ao enviar e-mail para %s: %s", target, message)

        worker.signals.result.connect(_on_result)
        worker.signals.error.connect(_on_error)
        self._thread_pool.start(worker)

    # ============================
    # DB: registrar pagamento
    # ============================
    def _registrar_pagamento_worker(self, dados: dict) -> dict:
        tipo_pagador = str((dados or {}).get("tipo_pagador") or "cliente").strip().lower()
        if tipo_pagador not in {"cliente", "empresa"}:
            tipo_pagador = "cliente"

        ok, msg = pagamento_controller.registrar_pagamento(dict(dados or {}))
        return {
            "ok": bool(ok),
            "msg": str(msg or ""),
            "tipo_pagador": tipo_pagador,
        }

    def registrar_pagamento_no_banco(self, dados: dict):
        if self._pagamento_register_inflight:
            try:
                if hasattr(self.pagamento, "_show_message"):
                    self.pagamento._show_message("Aguarde o registro atual terminar.", ok=True, ms=2200)
            except Exception:
                pass
            return

        self._pagamento_register_inflight = True
        try:
            if hasattr(self.pagamento, "_set_loading"):
                self.pagamento._set_loading(True)
        except Exception:
            pass

        worker = _Worker(self._registrar_pagamento_worker, dict(dados or {}))
        self._pagamento_register_workers.append(worker)

        def _cleanup():
            self._pagamento_register_inflight = False
            try:
                self._pagamento_register_workers.remove(worker)
            except Exception:
                pass

        def _on_result(result: dict):
            _cleanup()
            ok = bool((result or {}).get("ok"))
            msg = str((result or {}).get("msg") or "")
            tipo_pagador = str((result or {}).get("tipo_pagador") or "cliente").strip().lower()

            self.pagamento.finish_register(ok, msg)

            if tipo_pagador == "cliente" and hasattr(self.listar, "reload"):
                self.listar.reload()
            if tipo_pagador == "empresa" and hasattr(self.listar_empresas, "reload"):
                self.listar_empresas.reload()
            if ok:
                self._invalidate_dashboard_cache()
            self.atualizar_dashboard_async()

        def _on_error(error_msg: str):
            _cleanup()
            self.pagamento.finish_register(False, f"Erro interno ao registrar pagamento: {error_msg}")

        worker.signals.result.connect(_on_result)
        worker.signals.error.connect(_on_error)
        self._thread_pool.start(worker)
